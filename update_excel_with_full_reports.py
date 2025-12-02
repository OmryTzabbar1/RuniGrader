#!/usr/bin/env python3
"""
Extract full assessment text from student PDF reports and update Excel.

Instead of just summary strengths/weaknesses, this extracts the complete
feedback text from each student's PDF report.
"""
import os
import sys
from pathlib import Path
import openpyxl
from PyPDF2 import PdfReader

def find_student_folder(submissions_dir, student_id):
    """Find student's submission folder."""
    submissions_path = Path(submissions_dir)

    for participant_dir in submissions_path.glob(f"Participant_{student_id}*"):
        if participant_dir.is_dir():
            return str(participant_dir)

    return None

def extract_full_report_from_pdf(pdf_path):
    """
    Extract the complete assessment text from PDF.
    Returns: (strengths_text, weaknesses_text)
    """
    try:
        reader = PdfReader(pdf_path)

        # The assessment feedback is typically on page 2 (index 1)
        # Extract all text from the PDF
        full_text = ""
        for page in reader.pages:
            full_text += page.extract_text() + "\n"

        # Find the strengths section
        strengths_text = ""
        if "Strengths" in full_text or "Key Strengths" in full_text:
            # Extract text between Strengths and Improvements sections
            if "Strengths" in full_text:
                start_marker = "Strengths"
            else:
                start_marker = "Key Strengths"

            start_idx = full_text.find(start_marker)
            if start_idx != -1:
                # Find the improvements section
                end_markers = ["Required Improvements", "Areas for Improvement", "Areas Needing Attention", "To Reach Even Higher"]
                end_idx = len(full_text)
                for marker in end_markers:
                    idx = full_text.find(marker, start_idx)
                    if idx != -1:
                        end_idx = min(end_idx, idx)

                strengths_section = full_text[start_idx:end_idx].strip()
                # Clean up the section
                strengths_section = strengths_section.replace(start_marker, "").strip()
                # Remove bullet points and extra whitespace
                lines = [line.strip() for line in strengths_section.split('\n') if line.strip() and line.strip() != '•']
                strengths_text = '\n'.join(lines)

        # Find the weaknesses/improvements section
        weaknesses_text = ""
        improvement_markers = ["Required Improvements", "Areas for Improvement", "Areas Needing Attention", "To Reach Even Higher"]
        for marker in improvement_markers:
            if marker in full_text:
                start_idx = full_text.find(marker)
                if start_idx != -1:
                    # Find the end (usually "Work on these areas" or "Assessed:")
                    end_markers = ["Work on these areas", "Assessed:", "This grade reflects"]
                    end_idx = len(full_text)
                    for end_marker in end_markers:
                        idx = full_text.find(end_marker, start_idx)
                        if idx != -1:
                            end_idx = min(end_idx, idx)

                    weaknesses_section = full_text[start_idx:end_idx].strip()
                    # Clean up the section
                    weaknesses_section = weaknesses_section.replace(marker, "").strip()
                    # Remove bullet points and extra whitespace
                    lines = [line.strip() for line in weaknesses_section.split('\n') if line.strip() and line.strip() != '•']
                    weaknesses_text = '\n'.join(lines)
                    break

        return strengths_text, weaknesses_text

    except Exception as e:
        print(f"[ERROR] Failed to extract from PDF: {e}")
        return None, None

def main():
    print("=" * 60)
    print("UPDATE EXCEL WITH FULL PDF REPORTS")
    print("=" * 60)

    excel_path = "Assignment1_Grading_Summary.xlsx"
    submissions_dir = "WorkSubmissions01"

    print(f"\n[LOAD] Opening Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Find column indices
    header_row = 1
    headers = {}
    for cell in ws[header_row]:
        if cell.value:
            headers[cell.value] = cell.column

    student_id_col = headers.get('Student ID')
    strengths_col = headers.get('Key Strengths')
    weaknesses_col = headers.get('Key Weaknesses')

    if not all([student_id_col, strengths_col, weaknesses_col]):
        print(f"[ERROR] Required columns not found")
        return 1

    print(f"[OK] Found columns: Student ID={student_id_col}, Strengths={strengths_col}, Weaknesses={weaknesses_col}")

    stats = {
        'processed': 0,
        'updated': 0,
        'failed': 0
    }

    print(f"\n[UPDATE] Processing students...")
    print("-" * 60)

    # Process each student row
    for row_num in range(header_row + 1, ws.max_row + 1):
        student_id_val = ws.cell(row_num, student_id_col).value
        if not student_id_val:
            continue

        student_id = str(student_id_val).strip()

        # Skip summary rows
        if not student_id.isdigit():
            continue

        # Find student folder
        student_folder = find_student_folder(submissions_dir, student_id)
        if not student_folder:
            print(f"[SKIP] Student {student_id}: Folder not found")
            stats['failed'] += 1
            continue

        # Find PDF report
        pdf_path = os.path.join(student_folder, f"Student_Grade_Report_{student_id}.pdf")
        if not os.path.exists(pdf_path):
            print(f"[SKIP] Student {student_id}: PDF not found")
            stats['failed'] += 1
            continue

        # Extract full report text
        strengths_text, weaknesses_text = extract_full_report_from_pdf(pdf_path)

        if strengths_text is None:
            print(f"[FAIL] Student {student_id}: Could not extract from PDF")
            stats['failed'] += 1
            continue

        # Update Excel cells
        ws.cell(row_num, strengths_col).value = strengths_text
        ws.cell(row_num, weaknesses_col).value = weaknesses_text

        print(f"[OK] Student {student_id}: Updated with full report text")
        stats['updated'] += 1
        stats['processed'] += 1

    # Save Excel
    print(f"\n[SAVE] Saving updated Excel file...")
    wb.save(excel_path)

    # Print summary
    print("\n" + "=" * 60)
    print("UPDATE SUMMARY")
    print("=" * 60)
    print(f"\nStudents Processed: {stats['processed']}")
    print(f"Successfully Updated: {stats['updated']}")
    print(f"Failed: {stats['failed']}")

    print(f"\n[OK] Excel updated with full PDF report text!")
    print("=" * 60)

    return 0

if __name__ == '__main__':
    sys.exit(main())
