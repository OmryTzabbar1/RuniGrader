#!/usr/bin/env python3
"""
Add Self-Grade column to existing Assignment3_Grading_Summary.xlsx
Inserts column after "Generated Grade (/100)" and before "Weighted Grade"
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import shutil
from datetime import datetime

# Backup the original file
excel_path = "Assignment3_Grading_Summary.xlsx"
backup_path = excel_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy2(excel_path, backup_path)
print(f"[OK] Backup created: {backup_path}")

# Load workbook
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Find column indices from header row
header_row = 1
headers = {}
for cell in ws[header_row]:
    if cell.value:
        headers[cell.value] = cell.column

print(f"\nFound headers: {list(headers.keys())}")

# Find the Generated Grade column
generated_grade_col = None
for header, col in headers.items():
    if "Generated" in header and "Grade" in header:
        generated_grade_col = col
        break

if not generated_grade_col:
    print("ERROR: Could not find 'Generated Grade' column")
    exit(1)

print(f"Generated Grade column: {generated_grade_col}")

# Insert new column after Generated Grade
insert_col = generated_grade_col + 1
ws.insert_cols(insert_col)
print(f"[OK] Inserted new column at position {insert_col}")

# Style definitions
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add header for Self-Grade column
header_cell = ws.cell(header_row, insert_col)
header_cell.value = "Self-Grade"
header_cell.font = header_font
header_cell.fill = header_fill
header_cell.alignment = center_align
header_cell.border = thin_border

# Set column width
ws.column_dimensions[openpyxl.utils.get_column_letter(insert_col)].width = 12

# Format all data cells in the new column
for row_num in range(header_row + 1, ws.max_row + 1):
    cell = ws.cell(row_num, insert_col)
    cell.value = ""
    cell.alignment = center_align
    cell.fill = yellow_fill
    cell.border = thin_border

print(f"[OK] Formatted {ws.max_row - header_row} data cells")

# Save workbook
wb.save(excel_path)
print(f"\n[OK] Excel updated: {excel_path}")
print(f"  Self-Grade column inserted at position {insert_col}")
print(f"  All cells formatted with yellow highlighting (ready for data entry)")
