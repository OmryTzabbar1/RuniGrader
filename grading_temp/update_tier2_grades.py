#!/usr/bin/env python3
"""Update Excel with actual Tier 2 orchestrator grades."""

import openpyxl
import json
import os

# Load Excel
wb = openpyxl.load_workbook('grades_hw1.xlsx')
ws = wb.active

# Tier 2 orchestrator results
tier2_grades = {
    '38950': 39.0,
    '38951': 7.0,
    '38952': 42.0,
    '38953': 35.0,
    '38954': 23.5
}

# Update each Tier 2 student
for row in range(2, ws.max_row + 1):
    participant_id = str(ws.cell(row, 1).value)

    if participant_id in tier2_grades:
        final_grade = tier2_grades[participant_id]
        ws.cell(row, 9, value=final_grade)

        # Determine performance tier
        if final_grade >= 90:
            tier = "Excellence"
        elif final_grade >= 80:
            tier = "Good"
        elif final_grade >= 55:
            tier = "Potential"
        else:
            tier = "Below Standard"

        ws.cell(row, 10, value=tier)
        print(f"Updated {participant_id}: {final_grade} ({tier})")

# Save
wb.save('grades_hw1.xlsx')
print("\nExcel file updated successfully!")
