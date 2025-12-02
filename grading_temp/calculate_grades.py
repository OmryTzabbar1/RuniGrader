#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Calculate grades for all students"""

import openpyxl
import os
from openpyxl.styles import Font, PatternFill, Alignment

# Load Excel
wb = openpyxl.load_workbook('grades_hw1.xlsx')
ws = wb.active

# Add new columns if they don't exist
if ws.cell(1, 8).value != 'TRUE Count':
    ws.cell(1, 8, value='TRUE Count')
    ws.cell(1, 9, value='Final Grade')
    ws.cell(1, 10, value='Performance Tier')

    # Format headers
    for col in [8, 9, 10]:
        cell = ws.cell(1, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

print("Calculating grades for all students...")
print("="*60)

tier1_count = 0
tier2_count = 0

for row in range(2, ws.max_row + 1):
    participant_id = str(ws.cell(row, 1).value)
    self_grade = ws.cell(row, 6).value
    self_grade = int(self_grade) if self_grade else 0

    # Count TRUE criteria
    assessment_path = f"WorkSubmissions01/Participant_{participant_id}_assignsubmission_file/repo_assessment.md"
    true_count = 0

    if os.path.exists(assessment_path):
        with open(assessment_path, 'r', encoding='utf-8') as f:
            content = f.read()
            true_count = content.count('| TRUE |')

    # Write TRUE count
    ws.cell(row, 8, value=true_count)

    # Calculate grade based on tier
    if self_grade < 80:
        # Tier 1: Simple grading
        final_grade = (true_count / 22) * 100
        ws.cell(row, 9, value=round(final_grade, 1))

        # Determine performance tier
        if final_grade >= 90:
            tier = "Excellence"
        elif final_grade >= 80:
            tier = "Good"
        elif final_grade >= 55:
            tier = "Potential"
        else:
            tier = "Below Standard"

        ws.cell(row, 10, value=tier)
        tier1_count += 1
        print(f"[Tier 1] {participant_id}: {true_count}/22 = {final_grade:.1f}% ({tier})")
    else:
        # Tier 2: Will need 10 skills assessment
        ws.cell(row, 9, value="PENDING - Tier 2")
        ws.cell(row, 10, value="Tier 2 - Skills Needed")
        tier2_count += 1
        print(f"[Tier 2] {participant_id}: Self={self_grade}, TRUE={true_count}/22 - NEEDS 10 SKILLS")

# Save workbook
wb.save('grades_hw1.xlsx')

print("="*60)
print(f"\nGrading Summary:")
print(f"  Tier 1 (Simple): {tier1_count} students - COMPLETE")
print(f"  Tier 2 (Skills): {tier2_count} students - PENDING")
print(f"\nExcel updated: grades_hw1.xlsx")
