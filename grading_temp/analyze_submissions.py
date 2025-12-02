#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Analyze WorkSubmissions01 for grading"""

import openpyxl
import os
import sys

# Set UTF-8 encoding
if sys.platform == 'win32':
    import codecs
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

# Load Excel
wb = openpyxl.load_workbook('grades_hw1.xlsx')
ws = wb.active

print("WorkSubmissions01 Grading Analysis")
print("="*60)

students = []
for row in range(2, ws.max_row + 1):
    participant_id = str(ws.cell(row, 1).value)
    self_grade = ws.cell(row, 6).value

    # Count TRUE criteria
    assessment_path = f"WorkSubmissions01/Participant_{participant_id}_assignsubmission_file/repo_assessment.md"
    true_count = 0

    if os.path.exists(assessment_path):
        with open(assessment_path, 'r', encoding='utf-8') as f:
            content = f.read()
            true_count = content.count('| TRUE |')

    students.append({
        'id': participant_id,
        'self_grade': int(self_grade) if self_grade else 0,
        'true_count': true_count
    })

# Analyze tiers
tier1 = [s for s in students if s['self_grade'] < 80]
tier2 = [s for s in students if s['self_grade'] >= 80]

print(f"\nTotal Students: {len(students)}")
print(f"\nTier Breakdown:")
print(f"  Tier 1 (Self-Grade < 80): {len(tier1)} students")
print(f"  Tier 2 (Self-Grade >= 80): {len(tier2)} students")

print(f"\nTRUE Count Statistics:")
avg_true = sum(s['true_count'] for s in students) / len(students)
print(f"  Average: {avg_true:.1f}/22")
print(f"  Range: {min(s['true_count'] for s in students)}-{max(s['true_count'] for s in students)}/22")

print(f"\nTier 1 Students (Self-Grade < 80):")
for s in tier1:
    simple_grade = (s['true_count'] / 22) * 100
    print(f"  {s['id']}: Self={s['self_grade']}, TRUE={s['true_count']}/22, Final={simple_grade:.1f}%")

print(f"\nTier 2 Students (Self-Grade >= 80) - Need 10 Skills Assessment:")
for s in tier2:
    print(f"  {s['id']}: Self={s['self_grade']}, TRUE={s['true_count']}/22")

print("\n" + "="*60)

# Save analysis to file
with open('grading_analysis.txt', 'w', encoding='utf-8') as f:
    f.write(f"Total: {len(students)} students\n")
    f.write(f"Tier 1: {len(tier1)} students\n")
    f.write(f"Tier 2: {len(tier2)} students\n")
    f.write(f"\nTier 2 IDs:\n")
    for s in tier2:
        f.write(f"{s['id']}\n")

print("\nAnalysis saved to grading_analysis.txt")
