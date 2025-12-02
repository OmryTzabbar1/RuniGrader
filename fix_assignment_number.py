#!/usr/bin/env python3
"""
Fix assignment number from Assignment 3 to Assignment 1 everywhere.
1. Rename Excel file
2. Regenerate all PDF reports with correct assignment number
3. Regenerate merged PDFs
"""
import os
import sys
import shutil
from pathlib import Path

# Import the grade report generator
sys.path.append('.claude/skills/grade-report-generator')
from generate_student_report import create_student_report

# Import PDF merger
sys.path.append('.claude/skills/pdf-merger')
from merge_pdfs import find_pdfs, merge_pdfs

import openpyxl

def rename_excel_file():
    """Rename Assignment3_Grading_Summary.xlsx to Assignment1_Grading_Summary.xlsx"""
    old_name = "Assignment3_Grading_Summary.xlsx"
    new_name = "Assignment1_Grading_Summary.xlsx"

    if os.path.exists(old_name):
        if os.path.exists(new_name):
            print(f"[WARN] {new_name} already exists, skipping rename")
            return new_name
        try:
            os.rename(old_name, new_name)
            print(f"[OK] Renamed {old_name} -> {new_name}")
            return new_name
        except Exception as e:
            print(f"[ERROR] Could not rename Excel file: {e}")
            print(f"[ERROR] Please close the Excel file and run this script again")
            return None
    elif os.path.exists(new_name):
        print(f"[OK] {new_name} already exists")
        return new_name
    else:
        print(f"[ERROR] Neither {old_name} nor {new_name} found!")
        return None

def load_student_data(excel_path):
    """Load student data from Excel including weighted grades."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Find column indices
    header_row = 1
    headers = {}
    for cell in ws[header_row]:
        if cell.value:
            headers[cell.value] = cell.column

    student_id_col = headers.get('Student ID')
    team_col = headers.get('Team Name')
    base_grade_col = None
    weighted_grade_col = None
    strengths_col = headers.get('Key Strengths')
    weaknesses_col = headers.get('Key Weaknesses')

    for header, col in headers.items():
        if "Generated" in header and "Grade" in header:
            base_grade_col = col
        elif "Weighted" in header and "Grade" in header:
            weighted_grade_col = col

    if not all([student_id_col, team_col, base_grade_col, weighted_grade_col, strengths_col, weaknesses_col]):
        print("ERROR: Required columns not found in Excel")
        print(f"Found headers: {list(headers.keys())}")
        return []

    # Load student data
    students = []
    for row_num in range(header_row + 1, ws.max_row + 1):
        student_id_val = ws.cell(row_num, student_id_col).value
        if not student_id_val:
            continue

        student_id = str(student_id_val).strip()
        team = ws.cell(row_num, team_col).value
        base_grade = ws.cell(row_num, base_grade_col).value
        weighted_grade = ws.cell(row_num, weighted_grade_col).value
        strengths = ws.cell(row_num, strengths_col).value
        weaknesses = ws.cell(row_num, weaknesses_col).value

        if not weighted_grade:
            continue

        # Split strengths and weaknesses into lists (comma-separated)
        strengths_list = [s.strip() for s in str(strengths).split(',') if s.strip()] if strengths else []
        weaknesses_list = [w.strip() for w in str(weaknesses).split(',') if w.strip()] if weaknesses else []

        students.append({
            'student_id': student_id,
            'team': team or 'Unknown',
            'base_grade': float(base_grade) if base_grade else 0,
            'weighted_grade': float(weighted_grade),
            'strengths': strengths_list,
            'weaknesses': weaknesses_list
        })

    return students

def find_student_folder(submissions_dir, student_id):
    """Find student's submission folder."""
    submissions_path = Path(submissions_dir)

    for participant_dir in submissions_path.glob(f"Participant_{student_id}*"):
        if participant_dir.is_dir():
            return str(participant_dir)

    return None

def find_repository_url(student_folder):
    """Try to find repository URL from submission_info.xlsx."""
    try:
        excel_path = Path(student_folder) / "submission_info.xlsx"
        if excel_path.exists():
            wb = openpyxl.load_workbook(str(excel_path), data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=1, max_row=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if "github" in cell.value.lower() or "repository" in cell.value.lower():
                            # Next cell might have the URL
                            next_cell = ws.cell(cell.row, cell.column + 1).value
                            if next_cell and "github.com" in str(next_cell):
                                return str(next_cell)
    except:
        pass

    return "Repository not found"

def main():
    print("=" * 60)
    print("FIX ASSIGNMENT NUMBER: ASSIGNMENT 3 -> ASSIGNMENT 1")
    print("=" * 60)

    # Step 1: Rename Excel file
    print("\n[STEP 1] Renaming Excel file...")
    excel_path = rename_excel_file()
    if not excel_path:
        return 1

    # Step 2: Load student data
    print(f"\n[STEP 2] Loading student data from {excel_path}...")
    students = load_student_data(excel_path)
    print(f"[OK] Loaded {len(students)} students")

    # Step 3: Regenerate PDF reports with correct assignment number
    print(f"\n[STEP 3] Regenerating PDF reports with 'Assignment 1'...")
    submissions_dir = "WorkSubmissions01"

    stats = {
        'reports_regenerated': 0,
        'reports_failed': 0,
        'pdfs_merged': 0,
        'pdfs_failed': 0
    }

    for student in students:
        student_id = student['student_id']

        # Find student folder
        student_folder = find_student_folder(submissions_dir, student_id)
        if not student_folder:
            print(f"[SKIP] Student {student_id}: Folder not found")
            stats['reports_failed'] += 1
            continue

        # Find repository URL
        repository = find_repository_url(student_folder)

        # Generate report path
        report_path = os.path.join(student_folder, f"Student_Grade_Report_{student_id}.pdf")

        # Backup old report if exists
        if os.path.exists(report_path):
            backup_path = report_path.replace('.pdf', '_assignment3_backup.pdf')
            if not os.path.exists(backup_path):
                try:
                    shutil.copy2(report_path, backup_path)
                except:
                    pass

        try:
            # Generate new report with ASSIGNMENT 1 and WEIGHTED grade
            create_student_report(
                output_path=report_path,
                student_id=student_id,
                team=student['team'],
                grade=student['weighted_grade'],
                repository=repository,
                strengths=student['strengths'],
                improvements=student['weaknesses'],
                assignment="Assignment 1"  # CORRECT ASSIGNMENT NUMBER
            )

            print(f"[OK] Student {student_id}: Report regenerated (Grade: {student['weighted_grade']:.1f})")
            stats['reports_regenerated'] += 1

        except Exception as e:
            print(f"[FAIL] Student {student_id}: {e}")
            stats['reports_failed'] += 1
            continue

        # Step 4: Regenerate merged PDF
        try:
            submission_pdf, grade_report_pdf = find_pdfs(student_folder)

            if not submission_pdf or not grade_report_pdf:
                print(f"[SKIP] Student {student_id}: Missing PDFs for merge")
                stats['pdfs_failed'] += 1
                continue

            merged_path = os.path.join(student_folder, f"Student_{student_id}_Complete_Submission.pdf")

            # Backup old merged PDF if exists
            if os.path.exists(merged_path):
                backup_path = merged_path.replace('.pdf', '_assignment3_backup.pdf')
                if not os.path.exists(backup_path):
                    try:
                        shutil.copy2(merged_path, backup_path)
                    except:
                        pass

            success, page_info, error = merge_pdfs(submission_pdf, grade_report_pdf, merged_path)

            if success:
                stats['pdfs_merged'] += 1
            else:
                print(f"[WARN] Student {student_id}: Merge failed - {error}")
                stats['pdfs_failed'] += 1

        except Exception as e:
            print(f"[WARN] Student {student_id}: Could not merge - {e}")
            stats['pdfs_failed'] += 1

    # Print summary
    print("\n" + "=" * 60)
    print("FIX SUMMARY")
    print("=" * 60)
    print(f"\nReports Regenerated: {stats['reports_regenerated']}")
    print(f"Reports Failed: {stats['reports_failed']}")
    print(f"PDFs Merged: {stats['pdfs_merged']}")
    print(f"PDFs Failed: {stats['pdfs_failed']}")
    print("\n[OK] All fixes complete!")
    print("     - Excel renamed to Assignment1_Grading_Summary.xlsx")
    print("     - PDF reports now say 'Assignment 1'")
    print("     - Merged PDFs updated with corrected reports")
    print("     - Old files backed up as *_assignment3_backup.pdf")
    print("\n" + "=" * 60)

    return 0

if __name__ == '__main__':
    sys.exit(main())
