#!/usr/bin/env python3
"""
Regenerate all student grade reports with weighted (post-curve) grades instead of base grades.

Reads from Assignment3_Grading_Summary.xlsx and regenerates PDF reports.
"""
import os
import sys
import openpyxl
from pathlib import Path

# Import the grade report generator
sys.path.append('.claude/skills/grade-report-generator')
from generate_student_report import create_student_report

def load_student_data(excel_path):
    """Load student data from Excel including weighted grades."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Find column indices
    header_row = 1
    headers = {}
    for cell in ws[header_row]:
        if cell.value:
            headers[cell.value] = cell.column

    student_id_col = headers.get('Student ID')
    team_col = headers.get('Team Name')
    base_grade_col = None
    weighted_grade_col = None
    strengths_col = headers.get('Key Strengths')
    weaknesses_col = headers.get('Key Weaknesses')

    for header, col in headers.items():
        if "Generated" in header and "Grade" in header:
            base_grade_col = col
        elif "Weighted" in header and "Grade" in header:
            weighted_grade_col = col

    if not all([student_id_col, team_col, base_grade_col, weighted_grade_col, strengths_col, weaknesses_col]):
        print("ERROR: Required columns not found in Excel")
        print(f"Found headers: {list(headers.keys())}")
        return []

    print(f"[OK] Found columns:")
    print(f"  Student ID: {student_id_col}")
    print(f"  Team Name: {team_col}")
    print(f"  Base Grade: {base_grade_col}")
    print(f"  Weighted Grade: {weighted_grade_col}")
    print(f"  Strengths: {strengths_col}")
    print(f"  Weaknesses: {weaknesses_col}")

    # Load student data
    students = []
    for row_num in range(header_row + 1, ws.max_row + 1):
        student_id_val = ws.cell(row_num, student_id_col).value
        if not student_id_val:
            continue

        student_id = str(student_id_val).strip()
        team = ws.cell(row_num, team_col).value
        base_grade = ws.cell(row_num, base_grade_col).value
        weighted_grade = ws.cell(row_num, weighted_grade_col).value
        strengths = ws.cell(row_num, strengths_col).value
        weaknesses = ws.cell(row_num, weaknesses_col).value

        if not weighted_grade:
            print(f"[SKIP] Student {student_id}: No weighted grade")
            continue

        students.append({
            'student_id': student_id,
            'team': team or 'Unknown',
            'base_grade': float(base_grade) if base_grade else 0,
            'weighted_grade': float(weighted_grade),
            'strengths': strengths or 'Not specified',
            'weaknesses': weaknesses or 'Not specified'
        })

    return students

def find_student_folder(submissions_dir, student_id):
    """Find student's submission folder."""
    submissions_path = Path(submissions_dir)

    for participant_dir in submissions_path.glob(f"Participant_{student_id}*"):
        if participant_dir.is_dir():
            return str(participant_dir)

    return None

def find_repository_url(student_folder):
    """Try to find repository URL from submission_info.xlsx."""
    try:
        import openpyxl
        excel_path = Path(student_folder) / "submission_info.xlsx"
        if excel_path.exists():
            wb = openpyxl.load_workbook(str(excel_path), data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=1, max_row=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if "github" in cell.value.lower() or "repository" in cell.value.lower():
                            # Next cell might have the URL
                            next_cell = ws.cell(cell.row, cell.column + 1).value
                            if next_cell and "github.com" in str(next_cell):
                                return str(next_cell)
    except:
        pass

    return "Repository not found"

def main():
    print("=" * 60)
    print("REGENERATE GRADE REPORTS WITH WEIGHTED GRADES")
    print("=" * 60)

    # Load Excel data
    excel_path = "Assignment3_Grading_Summary.xlsx"
    print(f"\n[LOAD] Reading student data from: {excel_path}")
    students = load_student_data(excel_path)
    print(f"[OK] Loaded {len(students)} students with weighted grades")

    # Submissions directory
    submissions_dir = "WorkSubmissions01"

    # Statistics
    stats = {
        'processed': 0,
        'successful': 0,
        'failed': 0
    }

    print(f"\n[GENERATE] Regenerating grade reports...")
    print("-" * 60)

    for student in students:
        student_id = student['student_id']

        # Find student folder
        student_folder = find_student_folder(submissions_dir, student_id)
        if not student_folder:
            print(f"[FAIL] Student {student_id}: Folder not found")
            stats['failed'] += 1
            continue

        # Find repository URL
        repository = find_repository_url(student_folder)

        # Generate output path
        output_path = os.path.join(student_folder, f"Student_Grade_Report_{student_id}.pdf")

        # Create backup of old report if exists
        if os.path.exists(output_path):
            backup_path = output_path.replace('.pdf', '_old.pdf')
            if os.path.exists(backup_path):
                os.remove(backup_path)
            os.rename(output_path, backup_path)

        try:
            # Generate new report with WEIGHTED grade
            create_student_report(
                output_path=output_path,
                student_id=student_id,
                team=student['team'],
                grade=student['weighted_grade'],  # Use weighted grade!
                repository=repository,
                strengths=student['strengths'],
                improvements=student['weaknesses'],
                assignment="Assignment 3"
            )

            base = student['base_grade']
            weighted = student['weighted_grade']
            diff = base - weighted

            print(f"[OK] Student {student_id}: Base={base:.1f} -> Weighted={weighted:.1f} (penalty={diff:.1f})")
            stats['successful'] += 1

        except Exception as e:
            print(f"[FAIL] Student {student_id}: {e}")
            stats['failed'] += 1

        stats['processed'] += 1

    # Print summary
    print("\n" + "=" * 60)
    print("REGENERATION SUMMARY")
    print("=" * 60)
    print(f"\nStudents Processed: {stats['processed']}")
    print(f"Successful: {stats['successful']} ({stats['successful']/stats['processed']*100:.1f}%)")
    print(f"Failed: {stats['failed']} ({stats['failed']/stats['processed']*100:.1f}%)")

    print("\n[OK] All grade reports regenerated with weighted (post-curve) grades!")
    print("     Old reports backed up as *_old.pdf")
    print("\n" + "=" * 60)

    return 0

if __name__ == '__main__':
    sys.exit(main())
