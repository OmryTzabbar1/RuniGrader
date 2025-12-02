#!/usr/bin/env python3
"""
Calculate weighted grades using exponential self-grading penalty formula
"""
import os
import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import shutil
from datetime import datetime

# Constants from grading formula
SCALE_COEFFICIENT_A = 0.086603
SCALE_EXPONENT_B = 0.027465

def calculate_scale(self_grade: int) -> float:
    """Calculate exponential scale multiplier."""
    return SCALE_COEFFICIENT_A * math.exp(SCALE_EXPONENT_B * self_grade)

def calculate_weighted_grade(self_grade: int, base_grade: float) -> dict:
    """
    Calculate weighted grade with penalty for overconfidence.

    Returns dict with: weighted_grade, penalty, scale, difference
    """
    # Calculate scale
    scale = calculate_scale(self_grade)

    # Calculate difference
    difference = self_grade - base_grade

    # Apply penalty only if overestimated
    if difference > 0:
        penalty = difference * scale
        weighted_grade = max(0, base_grade - penalty)
    else:
        # Reward humility - no penalty
        penalty = 0
        weighted_grade = base_grade

    return {
        'weighted_grade': weighted_grade,
        'penalty': penalty,
        'scale': scale,
        'difference': difference
    }

# Main execution
print("=" * 60)
print("WEIGHTED GRADE CALCULATION")
print("=" * 60)

# Backup Excel
excel_path = "Assignment3_Grading_Summary.xlsx"
backup_path = excel_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy2(excel_path, backup_path)
print(f"\n[OK] Backup created: {backup_path}")

# Load workbook
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Find column indices
header_row = 1
headers = {}
for cell in ws[header_row]:
    if cell.value:
        headers[cell.value] = cell.column

student_id_col = headers.get('Student ID')
self_grade_col = headers.get('Self-Grade')
weighted_grade_col = None
generated_grade_col = None

for header, col in headers.items():
    if "Generated" in header and "Grade" in header:
        generated_grade_col = col
    elif "Weighted" in header and "Grade" in header:
        weighted_grade_col = col

if not all([student_id_col, self_grade_col, generated_grade_col, weighted_grade_col]):
    print("ERROR: Required columns not found")
    print(f"Found headers: {list(headers.keys())}")
    exit(1)

print(f"\n[OK] Found columns:")
print(f"  Student ID: {student_id_col}")
print(f"  Self-Grade: {self_grade_col}")
print(f"  Generated Grade: {generated_grade_col}")
print(f"  Weighted Grade: {weighted_grade_col}")

# Statistics
stats = {
    'processed': 0,
    'no_penalty': 0,
    'small_penalty': 0,
    'medium_penalty': 0,
    'large_penalty': 0,
    'penalties': []
}

# Grade color fills
grade_a = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
grade_b = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
grade_c = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
grade_d = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
grade_f = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

print(f"\n[CALCULATE] Processing students...")
print("-" * 60)

# Process each student row
for row_num in range(header_row + 1, ws.max_row + 1):
    student_id_cell = ws.cell(row_num, student_id_col).value
    if not student_id_cell:
        continue

    student_id = str(student_id_cell).strip()

    self_grade_val = ws.cell(row_num, self_grade_col).value
    base_grade_val = ws.cell(row_num, generated_grade_col).value

    if not self_grade_val or not base_grade_val:
        print(f"[SKIP] Student {student_id}: Missing self-grade or base grade")
        continue

    self_grade = int(float(self_grade_val))
    base_grade = float(base_grade_val)

    # Calculate weighted grade
    result = calculate_weighted_grade(self_grade, base_grade)

    weighted_grade = result['weighted_grade']
    penalty = result['penalty']

    # Track statistics
    if penalty == 0:
        stats['no_penalty'] += 1
        status = "HUMBLE/ACCURATE"
    elif penalty <= 5:
        stats['small_penalty'] += 1
        status = "SLIGHTLY OPTIMISTIC"
    elif penalty <= 15:
        stats['medium_penalty'] += 1
        status = "OVERCONFIDENT"
    else:
        stats['large_penalty'] += 1
        status = "VERY OVERCONFIDENT"

    stats['penalties'].append({
        'student_id': student_id,
        'self_grade': self_grade,
        'base_grade': base_grade,
        'penalty': penalty,
        'weighted_grade': weighted_grade,
        'status': status
    })

    print(f"[{status:20s}] Student {student_id}: Self={self_grade}, Base={base_grade:.0f}, Penalty={penalty:.2f}, Weighted={weighted_grade:.2f}")

    # Update Excel - Weighted Grade column
    weighted_cell = ws.cell(row_num, weighted_grade_col)
    weighted_cell.value = round(weighted_grade, 2)
    weighted_cell.alignment = Alignment(horizontal="center", vertical="center")
    weighted_cell.font = Font(bold=True, size=11)

    # Color code weighted grade
    if weighted_grade >= 90:
        weighted_cell.fill = grade_a
    elif weighted_grade >= 80:
        weighted_cell.fill = grade_b
    elif weighted_grade >= 70:
        weighted_cell.fill = grade_c
    elif weighted_grade >= 60:
        weighted_cell.fill = grade_d
    else:
        weighted_cell.fill = grade_f

    stats['processed'] += 1

# Save workbook
wb.save(excel_path)

# Print summary
print("\n" + "=" * 60)
print("SUMMARY REPORT")
print("=" * 60)
print(f"\nStudents Processed: {stats['processed']}")

print(f"\nPenalty Statistics:")
print(f"  No Penalty (accurate/humble): {stats['no_penalty']} ({stats['no_penalty']/stats['processed']*100:.1f}%)")
print(f"  Small Penalty (1-5 points): {stats['small_penalty']} ({stats['small_penalty']/stats['processed']*100:.1f}%)")
print(f"  Medium Penalty (6-15 points): {stats['medium_penalty']} ({stats['medium_penalty']/stats['processed']*100:.1f}%)")
print(f"  Large Penalty (>15 points): {stats['large_penalty']} ({stats['large_penalty']/stats['processed']*100:.1f}%)")

# Calculate averages
penalties = stats['penalties']
if penalties:
    avg_base = sum(p['base_grade'] for p in penalties) / len(penalties)
    avg_weighted = sum(p['weighted_grade'] for p in penalties) / len(penalties)
    avg_penalty = sum(p['penalty'] for p in penalties) / len(penalties)

    print(f"\nAverage Base Grade: {avg_base:.2f}")
    print(f"Average Weighted Grade: {avg_weighted:.2f}")
    print(f"Average Penalty: {avg_penalty:.2f} points")

    # Top self-assessors (most accurate)
    sorted_by_accuracy = sorted(penalties, key=lambda p: abs(p['self_grade'] - p['base_grade']))
    print(f"\nTop 5 Self-Assessors (most accurate):")
    for i, p in enumerate(sorted_by_accuracy[:5], 1):
        diff = abs(p['self_grade'] - p['base_grade'])
        print(f"  {i}. Student {p['student_id']}: Self={p['self_grade']}, Base={p['base_grade']:.0f}, Diff={diff:.1f}")

    # Bottom self-assessors (most overconfident)
    overconfident = [p for p in penalties if p['penalty'] > 0]
    if overconfident:
        sorted_by_overconfidence = sorted(overconfident, key=lambda p: p['penalty'], reverse=True)
        print(f"\nTop 5 Most Overconfident:")
        for i, p in enumerate(sorted_by_overconfidence[:5], 1):
            print(f"  {i}. Student {p['student_id']}: Self={p['self_grade']}, Base={p['base_grade']:.0f}, Penalty={p['penalty']:.2f}, Weighted={p['weighted_grade']:.2f}")

print(f"\n[OK] Excel updated: {excel_path}")
print(f"  Weighted Grade column populated for {stats['processed']} students")
print(f"  Backup: {backup_path}")
print("\n" + "=" * 60)
