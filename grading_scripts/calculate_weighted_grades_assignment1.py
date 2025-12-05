#!/usr/bin/env python3
"""
Calculate weighted grades for Assignment 1 using the exponential penalty formula.
Updates Column I with the weighted grade (what they would have gotten with penalty).
"""
import openpyxl
import json
import math
from pathlib import Path
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Constants from weighted-grade-calculator skill
SCALE_COEFFICIENT_A = 0.086603
SCALE_EXPONENT_B = 0.027465

def calculate_weighted_grade(self_grade, base_grade):
    """
    Calculate weighted grade using exponential penalty formula.

    Formula:
    1. scale = 0.086603 × e^(0.027465 × self_grade)
    2. if self_grade > base_grade:
           penalty = (self_grade - base_grade) × scale
           weighted = max(0, base_grade - penalty)
       else:
           weighted = base_grade  # No penalty for humility
    """
    # Calculate scale multiplier
    scale = SCALE_COEFFICIENT_A * math.exp(SCALE_EXPONENT_B * self_grade)

    # Apply penalty only if overestimated
    if self_grade > base_grade:
        difference = self_grade - base_grade
        penalty = difference * scale
        weighted_grade = max(0, base_grade - penalty)
    else:
        # Humble or accurate - no penalty
        weighted_grade = base_grade

    return round(weighted_grade, 2)

# Load Excel
wb = openpyxl.load_workbook('StudentGradesMoodleFormat/Assignment1_Moodle_Grades_UPDATED.xlsx')
ws = wb.active

assessments_dir = Path("assessments_tier2_assignment1")

print("=== Calculating Weighted Grades for Assignment 1 ===\n")
print("Formula: Exponential penalty for overconfidence\n")

updates = 0
calculations = []

# Process each row
for row_idx in range(2, ws.max_row + 1):
    identifier = ws[f'A{row_idx}'].value

    if not identifier or ':' not in str(identifier):
        continue

    # Extract student ID
    student_id = identifier.split(':')[1].strip()

    # Get self-grade from Column C
    self_grade = ws[f'C{row_idx}'].value

    # Get assessment data for base grade
    json_file = assessments_dir / f"tier2_assessment_{student_id}.json"

    if json_file.exists() and self_grade is not None:
        with open(json_file, 'r', encoding='utf-8') as f:
            assessment = json.load(f)

        # Base grade is the total_score from Tier 2 assessment
        base_grade = assessment.get('total_score', 0)

        # Calculate weighted grade with penalty formula
        weighted_grade = calculate_weighted_grade(self_grade, base_grade)

        # Round up for display
        weighted_grade_display = math.ceil(weighted_grade)

        # Update Column I
        ws[f'I{row_idx}'] = weighted_grade_display

        penalty = base_grade - weighted_grade
        calculations.append({
            'student_id': student_id,
            'self_grade': self_grade,
            'base_grade': base_grade,
            'weighted_grade': weighted_grade,
            'weighted_display': weighted_grade_display,
            'penalty': penalty
        })

        penalty_str = f"-{penalty:.1f}" if penalty > 0 else "0"
        print(f"Student {student_id}: Self={self_grade:.0f}, Base={base_grade:.1f}, Weighted={weighted_grade:.1f} ({weighted_grade_display}), Penalty={penalty_str}")
        updates += 1

# Save new version
output_path = Path('StudentGradesMoodleFormat/Assignment1_Moodle_Grades_FINAL.xlsx')
wb.save(output_path)

print(f"\n=== Summary ===")
print(f"Students processed: {updates}")

# Statistics
penalties = [c['penalty'] for c in calculations]
no_penalty = sum(1 for p in penalties if p <= 0.1)
small_penalty = sum(1 for p in penalties if 0.1 < p <= 5)
medium_penalty = sum(1 for p in penalties if 5 < p <= 15)
large_penalty = sum(1 for p in penalties if p > 15)

print(f"\nPenalty Statistics:")
print(f"  No penalty (accurate/humble): {no_penalty} ({no_penalty/updates*100:.0f}%)")
print(f"  Small penalty (0-5 points): {small_penalty} ({small_penalty/updates*100:.0f}%)")
print(f"  Medium penalty (5-15 points): {medium_penalty} ({medium_penalty/updates*100:.0f}%)")
print(f"  Large penalty (>15 points): {large_penalty} ({large_penalty/updates*100:.0f}%)")

avg_penalty = sum(penalties) / len(penalties)
print(f"\nAverage penalty: {avg_penalty:.2f} points")

# Show top overestimators
sorted_calcs = sorted(calculations, key=lambda x: x['penalty'], reverse=True)
print(f"\nTop 5 Overestimators:")
for calc in sorted_calcs[:5]:
    print(f"  {calc['student_id']}: Self={calc['self_grade']:.0f}, Base={calc['base_grade']:.1f}, Penalty=-{calc['penalty']:.1f}")

print(f"\nFile saved: {output_path}")
print("Column I now contains weighted grades with penalty formula applied (rounded up)")
