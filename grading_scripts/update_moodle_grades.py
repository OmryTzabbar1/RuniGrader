#!/usr/bin/env python3
"""
Update Moodle grade Excel files for Assignment 1:
1. Replace Grade column with self-submitted grade
2. Move weighted grade (rounded up) to column I
3. Replace column H with detailed feedback report text
"""
import openpyxl
import json
import math
from pathlib import Path

def get_self_submitted_grade(student_id, submissions_dir):
    """Get self-submitted grade from submission_info.xlsx"""
    student_folder = submissions_dir / f"Participant_{student_id}_assignsubmission_file"
    submission_info = student_folder / "submission_info.xlsx"

    if not submission_info.exists():
        return None

    try:
        wb = openpyxl.load_workbook(submission_info)
        ws = wb.active

        # Find the self-submitted grade (usually in a cell labeled "Grade" or similar)
        for row in ws.iter_rows(min_row=1, max_row=10):
            for idx, cell in enumerate(row):
                if cell.value and "grade" in str(cell.value).lower():
                    # Grade is likely in the next cell or same row
                    if idx + 1 < len(row):
                        grade_value = row[idx + 1].value
                        if grade_value and isinstance(grade_value, (int, float)):
                            return float(grade_value)

        # Alternative: check specific cells
        for cell_ref in ['B2', 'C2', 'D2', 'B3', 'C3']:
            value = ws[cell_ref].value
            if value and isinstance(value, (int, float)) and 0 <= value <= 100:
                return float(value)

    except Exception as e:
        print(f"  Warning: Could not read self-grade for {student_id}: {e}")

    return None

def get_assessment_data(student_id, assessments_dir):
    """Get assessment data from JSON"""
    json_file = assessments_dir / f"tier2_assessment_{student_id}.json"

    if not json_file.exists():
        return None

    with open(json_file, 'r', encoding='utf-8') as f:
        return json.load(f)

def generate_feedback_text(assessment_data):
    """Generate detailed feedback text without scores"""
    total_score = assessment_data.get('total_score', 0)
    strengths = assessment_data.get('overall_assessment', {}).get('strengths', [])
    weaknesses = assessment_data.get('overall_assessment', {}).get('weaknesses', [])
    recommendations = assessment_data.get('recommended_actions', {})

    # Determine performance level
    if total_score >= 80:
        tone = "excellent work"
    elif total_score >= 70:
        tone = "strong foundation"
    elif total_score >= 55:
        tone = "good potential"
    else:
        tone = "developing skills"

    feedback = f"Your submission demonstrates {tone}. "

    # Add strengths
    if strengths and strengths[0] != "No skills with excellent scores (8+/10)":
        strength_names = [s.split('(')[0].strip() for s in strengths[:3]]
        if len(strength_names) > 0:
            feedback += f"Key strengths: {', '.join(strength_names)}. "

    # Add areas for improvement
    if weaknesses and weaknesses[0] != "No critical weaknesses (<5/10)":
        weakness_names = [w.split('(')[0].strip() for w in weaknesses[:3]]
        if len(weakness_names) > 0:
            feedback += f"Focus areas: {', '.join(weakness_names)}. "

    # Add immediate actions
    immediate = recommendations.get('immediate', [])
    if immediate and immediate[0] != "All critical areas addressed - focus on optimization":
        action_items = [a.split('(')[0].strip() for a in immediate[:2]]
        if len(action_items) > 0:
            feedback += f"Priority actions: {', '.join(action_items)}. "

    # Add encouraging note
    if total_score >= 80:
        feedback += "Continue building on your strong foundation."
    elif total_score >= 55:
        feedback += "With focused improvements in these areas, you can significantly strengthen your work."
    else:
        feedback += "Focus on the priority areas outlined in your detailed report."

    return feedback

def update_assignment1_grades():
    """Update Assignment1_Moodle_Grades.xlsx"""

    print("=" * 80)
    print("UPDATING ASSIGNMENT 1 MOODLE GRADES")
    print("=" * 80)

    excel_path = Path("StudentGradesMoodleFormat/Assignment1_Moodle_Grades.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    submissions_dir = Path("WorkSubmissions01")
    assessments_dir = Path("assessments_tier2_assignment1")

    # Opening message for column H
    opening_message = ("Thank you for your submission. For this first assignment I have decided to take "
                      "your self-submitted grade as your grade for this assignment since many students "
                      "over-estimated their grades, and suffered severe penalties. You can see the grade "
                      "you would have gotten on this assignment in the next column. Do not let this "
                      "discourage you, you have a lot of potential, and can perform strongly in the "
                      "future with more effort. ")

    # Find the header row and columns
    # Assuming row 1 is headers
    # Need to identify: Grade column, Feedback column

    print("\nProcessing students...")

    updates = {
        'processed': 0,
        'self_grades_found': 0,
        'self_grades_missing': 0,
        'feedback_added': 0
    }

    # Process each row (skip header)
    for row_idx in range(2, ws.max_row + 1):
        # Column A typically has student identifier
        identifier = ws[f'A{row_idx}'].value

        if not identifier:
            continue

        # Extract student ID (assuming it's in the identifier or nearby columns)
        student_id = None
        for col_letter in ['A', 'B', 'C', 'D']:
            value = ws[f'{col_letter}{row_idx}'].value
            if value and isinstance(value, (int, str)):
                value_str = str(value)
                if value_str.isdigit() and len(value_str) == 5:
                    student_id = value_str
                    break

        if not student_id:
            print(f"Row {row_idx}: Could not find student ID")
            continue

        updates['processed'] += 1

        # Get self-submitted grade
        self_grade = get_self_submitted_grade(student_id, submissions_dir)

        # Get assessment data
        assessment = get_assessment_data(student_id, assessments_dir)

        if assessment:
            weighted_grade = assessment.get('final_grade', 0)

            # Column G or H is typically Grade - check both
            # We'll update column G with self-grade
            if self_grade is not None:
                ws[f'G{row_idx}'] = self_grade
                updates['self_grades_found'] += 1
            else:
                # If no self-grade found, use weighted grade
                ws[f'G{row_idx}'] = math.ceil(weighted_grade)
                updates['self_grades_missing'] += 1

            # Column I: Weighted grade (rounded up)
            ws[f'I{row_idx}'] = math.ceil(weighted_grade)

            # Column H: Feedback with opening message
            feedback_text = generate_feedback_text(assessment)
            full_feedback = opening_message + feedback_text
            ws[f'H{row_idx}'] = full_feedback
            updates['feedback_added'] += 1

            print(f"Row {row_idx}: Student {student_id} - Self: {self_grade}, Weighted: {math.ceil(weighted_grade)}")

    # Update column headers
    ws['G1'] = 'Grade (Self-Submitted)'
    ws['H1'] = 'Feedback comments'
    ws['I1'] = 'Calculated Grade (Tier 2)'

    # Save workbook
    output_path = Path("StudentGradesMoodleFormat/Assignment1_Moodle_Grades_Updated.xlsx")
    wb.save(output_path)

    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Students processed: {updates['processed']}")
    print(f"Self-grades found: {updates['self_grades_found']}")
    print(f"Self-grades missing (used weighted): {updates['self_grades_missing']}")
    print(f"Feedback added: {updates['feedback_added']}")
    print(f"\nOutput saved to: {output_path}")
    print("=" * 80)

if __name__ == '__main__':
    update_assignment1_grades()
