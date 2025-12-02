#!/usr/bin/env python3
"""
Populate Self-Grade column by extracting self-proclaimed grades from student PDFs
"""
import os
import re
import openpyxl
from pathlib import Path
import shutil
from datetime import datetime

try:
    from PyPDF2 import PdfReader
except ImportError:
    print("ERROR: PyPDF2 not installed. Run: pip install PyPDF2")
    exit(1)

# Constants
MIN_SELF_GRADE = 60
MAX_SELF_GRADE = 100

def extract_self_grade_from_pdf(pdf_path):
    """
    Extract self-proclaimed grade from PDF.
    Returns integer grade or None if not found.
    """
    if not os.path.exists(pdf_path):
        return None

    try:
        reader = PdfReader(pdf_path)

        # Extract all text
        text = ""
        for page in reader.pages:
            text += page.extract_text()

        # Search patterns
        patterns = [
            r"Self[- ]?Grade:?\s*(\d+)",       # "Self-Grade: 85"
            r"I estimate:?\s*(\d+)%?",         # "I estimate: 90%"
            r"My grade:?\s*(\d+)",              # "My grade: 80"
            r"Self[- ]?assessment:?\s*(\d+)",   # "Self-assessment: 88"
            r"Expected grade:?\s*(\d+)",        # "Expected grade: 92"
            r"Predicted grade:?\s*(\d+)"        # "Predicted grade: 87"
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                grade = int(match.group(1))
                # Validate range
                if MIN_SELF_GRADE <= grade <= MAX_SELF_GRADE:
                    return grade

        return None

    except Exception as e:
        print(f"  Warning: Could not extract from {pdf_path}: {e}")
        return None

def find_student_pdfs(submissions_dir):
    """Find all student submission PDFs."""
    pdf_map = {}

    submissions_path = Path(submissions_dir)
    if not submissions_path.exists():
        print(f"ERROR: Submissions directory not found: {submissions_dir}")
        return pdf_map

    for participant_dir in submissions_path.glob("Participant_*"):
        if not participant_dir.is_dir():
            continue

        # Extract student ID
        match = re.search(r"Participant_(\d+)", participant_dir.name)
        if not match:
            continue

        student_id = match.group(1)

        # Find submission PDF (not grade report)
        for pdf_file in participant_dir.glob("*.pdf"):
            # Skip generated grade reports
            if "Student_Grade_Report" in pdf_file.name:
                continue

            pdf_map[student_id] = str(pdf_file)
            break

    return pdf_map

# Main execution
print("=" * 60)
print("SELF-GRADE EXTRACTION FROM STUDENT PDFS")
print("=" * 60)

# Backup Excel
excel_path = "Assignment3_Grading_Summary.xlsx"
backup_path = excel_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy2(excel_path, backup_path)
print(f"\n[OK] Backup created: {backup_path}")

# Load workbook
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Find column indices
header_row = 1
headers = {}
for cell in ws[header_row]:
    if cell.value:
        headers[cell.value] = cell.column

student_id_col = headers.get('Student ID')
self_grade_col = headers.get('Self-Grade')
generated_grade_col = None

for header, col in headers.items():
    if "Generated" in header and "Grade" in header:
        generated_grade_col = col
        break

if not all([student_id_col, self_grade_col, generated_grade_col]):
    print("ERROR: Required columns not found")
    print(f"Found headers: {list(headers.keys())}")
    exit(1)

print(f"\n[OK] Found columns:")
print(f"  Student ID: {student_id_col}")
print(f"  Self-Grade: {self_grade_col}")
print(f"  Generated Grade: {generated_grade_col}")

# Find student PDFs
submissions_dir = r"C:\Users\Guest1\CoOp\Runi\WorkSubmissions01"
print(f"\n[SCAN] Looking for student PDFs in: {submissions_dir}")
pdf_map = find_student_pdfs(submissions_dir)
print(f"[OK] Found {len(pdf_map)} student submission PDFs")

# Statistics
stats = {
    'processed': 0,
    'found': 0,
    'missing': 0,
    'using_base': 0
}

results = []

print(f"\n[EXTRACT] Processing students...")
print("-" * 60)

# Process each student row
for row_num in range(header_row + 1, ws.max_row + 1):
    student_id_cell = ws.cell(row_num, student_id_col).value
    if not student_id_cell:
        continue

    student_id = str(student_id_cell).strip()
    generated_grade = ws.cell(row_num, generated_grade_col).value

    if not generated_grade:
        print(f"[SKIP] Student {student_id}: No generated grade")
        continue

    base_grade = int(float(generated_grade))

    # Try to extract self-grade from PDF
    self_grade = None

    if student_id in pdf_map:
        self_grade = extract_self_grade_from_pdf(pdf_map[student_id])

        if self_grade:
            print(f"[FOUND] Student {student_id}: Self-grade {self_grade} extracted")
            stats['found'] += 1
            results.append({
                'student_id': student_id,
                'self_grade': self_grade,
                'base_grade': base_grade,
                'source': 'PDF'
            })
        else:
            print(f"[NONE] Student {student_id}: No self-grade found, using base grade {base_grade}")
            self_grade = base_grade
            stats['missing'] += 1
            stats['using_base'] += 1
            results.append({
                'student_id': student_id,
                'self_grade': self_grade,
                'base_grade': base_grade,
                'source': 'Base Grade (no PDF self-grade)'
            })
    else:
        print(f"[NONE] Student {student_id}: No PDF found, using base grade {base_grade}")
        self_grade = base_grade
        stats['missing'] += 1
        stats['using_base'] += 1
        results.append({
            'student_id': student_id,
            'self_grade': self_grade,
            'base_grade': base_grade,
            'source': 'Base Grade (no PDF)'
        })

    # Update Excel
    ws.cell(row_num, self_grade_col).value = self_grade
    stats['processed'] += 1

# Save workbook
wb.save(excel_path)

# Print summary
print("\n" + "=" * 60)
print("SUMMARY REPORT")
print("=" * 60)
print(f"\nStudents Processed: {stats['processed']}")
print(f"Self-Grades Found in PDFs: {stats['found']} ({stats['found']/stats['processed']*100:.1f}%)")
print(f"Self-Grades Missing: {stats['missing']} ({stats['missing']/stats['processed']*100:.1f}%)")
print(f"Using Base Grade as Default: {stats['using_base']}")

print(f"\n[OK] Excel updated: {excel_path}")
print(f"  Self-Grade column populated for {stats['processed']} students")
print("\n" + "=" * 60)
