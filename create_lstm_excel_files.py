#!/usr/bin/env python3
"""
Create Excel files for LSTM assignment (WorkSubmissions02)
Follows the same format as WorkSubmissions01 Excel files
"""

import openpyxl
from openpyxl.styles import Font, Alignment
import json
import os

def create_excel_file(data, output_path):
    """Create Excel file with submission information."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submission Info"

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80

    # Header row
    ws['A1'] = 'Field'
    ws['A1'].font = Font(bold=True)
    ws['B1'] = 'Value'
    ws['B1'].font = Font(bold=True)

    # Data rows
    ws['A2'] = 'Participant ID'
    ws['B2'] = data['participant_id']

    ws['A3'] = 'Group Code'
    ws['B3'] = data['group_code']

    ws['A4'] = 'Student 1'
    ws['B4'] = data['student1']

    ws['A5'] = 'Student 2'
    ws['B5'] = data['student2']

    ws['A6'] = 'GitHub Repository'
    ws['B6'] = data['github_url']

    ws['A7'] = 'Suggested Grade'
    ws['B7'] = data['suggested_grade']

    ws['A8'] = 'PDF Filename'
    ws['B8'] = data['pdf_filename']

    # Save workbook
    wb.save(output_path)
    return True

def main():
    # Load extracted data
    with open(r'C:\Users\Guest1\CoOp\Runi\lstm_extracted_data.json', 'r', encoding='utf-8') as f:
        data = json.load(f)

    print(f'Creating Excel files for {len(data)} participants...\n')

    created_count = 0
    skipped_count = 0

    for entry in data:
        folder_path = entry['folder_path']
        excel_path = os.path.join(folder_path, 'submission_info.xlsx')

        # Prepare data for Excel
        excel_data = {
            'participant_id': entry['participant_id'],
            'group_code': entry['group_code'],
            'student1': f"{entry['student1_id']}" if entry['student1_id'] else '',
            'student2': f"{entry['student2_id']}" if entry['student2_id'] else '',
            'github_url': entry['github_url'],
            'suggested_grade': entry['suggested_grade'],
            'pdf_filename': entry['pdf_filename']
        }

        # Create Excel file
        try:
            create_excel_file(excel_data, excel_path)
            created_count += 1
            print(f"[OK] Created: {entry['participant_id']} - {excel_data['group_code'] or 'NO GROUP'}")
        except Exception as e:
            skipped_count += 1
            print(f"[FAIL] Failed: {entry['participant_id']} - {str(e)}")

    print(f'\n{"=" * 80}')
    print(f'Summary:')
    print(f'  Created: {created_count}')
    print(f'  Failed: {skipped_count}')
    print(f'  Total: {len(data)}')
    print(f'{"=" * 80}')

if __name__ == '__main__':
    main()
