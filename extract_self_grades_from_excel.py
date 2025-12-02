#!/usr/bin/env python3
"""
Extract self-grades from each student's submission_info.xlsx file
"""
import os
import openpyxl
from pathlib import Path
import shutil
from datetime import datetime

def find_student_excel_files(submissions_dir):
    """Find all student submission_info.xlsx files."""
    excel_map = {}

    submissions_path = Path(submissions_dir)
    if not submissions_path.exists():
        print(f"ERROR: Submissions directory not found: {submissions_dir}")
        return excel_map

    for participant_dir in submissions_path.glob("Participant_*"):
        if not participant_dir.is_dir():
            continue

        # Extract student ID from directory name
        import re
        match = re.search(r"Participant_(\d+)", participant_dir.name)
        if not match:
            continue

        student_id = match.group(1)

        # Look for submission_info.xlsx
        excel_file = participant_dir / "submission_info.xlsx"
        if excel_file.exists():
            excel_map[student_id] = str(excel_file)

    return excel_map

def extract_self_grade_from_excel(excel_path):
    """
    Extract self-grade from student's submission_info.xlsx
    Looks for "Suggested Grade" field
    Returns integer grade or None if not found.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        # Look for "Suggested Grade" in the spreadsheet
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Check for "Suggested Grade", "Self-Grade", etc.
                    cell_lower = cell.value.lower()
                    if ("suggest" in cell_lower and "grade" in cell_lower) or \
                       ("self" in cell_lower and "grade" in cell_lower):
                        # Found the label, get value from next column
                        value_cell = ws.cell(cell.row, cell.column + 1).value
                        if value_cell is not None:
                            try:
                                # Handle percentage strings or plain numbers
                                value_str = str(value_cell).strip().strip('%')
                                grade = int(float(value_str))
                                # Validate range
                                if 0 <= grade <= 100:
                                    return grade
                            except:
                                pass

        return None

    except Exception as e:
        print(f"  Warning: Could not read {excel_path}: {e}")
        return None

# Main execution
print("=" * 60)
print("SELF-GRADE EXTRACTION FROM STUDENT EXCEL FILES")
print("=" * 60)

# Backup Excel
excel_path = "Assignment3_Grading_Summary.xlsx"
backup_path = excel_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy2(excel_path, backup_path)
print(f"\n[OK] Backup created: {backup_path}")

# Load workbook
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Find column indices
header_row = 1
headers = {}
for cell in ws[header_row]:
    if cell.value:
        headers[cell.value] = cell.column

student_id_col = headers.get('Student ID')
self_grade_col = headers.get('Self-Grade')
generated_grade_col = None

for header, col in headers.items():
    if "Generated" in header and "Grade" in header:
        generated_grade_col = col
        break

if not all([student_id_col, self_grade_col, generated_grade_col]):
    print("ERROR: Required columns not found")
    print(f"Found headers: {list(headers.keys())}")
    exit(1)

print(f"\n[OK] Found columns:")
print(f"  Student ID: {student_id_col}")
print(f"  Self-Grade: {self_grade_col}")
print(f"  Generated Grade: {generated_grade_col}")

# Find student Excel files
submissions_dir = r"C:\Users\Guest1\CoOp\Runi\WorkSubmissions01"
print(f"\n[SCAN] Looking for submission_info.xlsx files in: {submissions_dir}")
excel_map = find_student_excel_files(submissions_dir)
print(f"[OK] Found {len(excel_map)} student Excel files")

# Statistics
stats = {
    'processed': 0,
    'found': 0,
    'missing': 0,
    'using_base': 0
}

results = []

print(f"\n[EXTRACT] Processing students...")
print("-" * 60)

# Process each student row
for row_num in range(header_row + 1, ws.max_row + 1):
    student_id_cell = ws.cell(row_num, student_id_col).value
    if not student_id_cell:
        continue

    student_id = str(student_id_cell).strip()
    generated_grade = ws.cell(row_num, generated_grade_col).value

    if not generated_grade:
        print(f"[SKIP] Student {student_id}: No generated grade")
        continue

    base_grade = int(float(generated_grade))

    # Try to extract self-grade from Excel
    self_grade = None

    if student_id in excel_map:
        self_grade = extract_self_grade_from_excel(excel_map[student_id])

        if self_grade:
            print(f"[FOUND] Student {student_id}: Self-grade {self_grade} from Excel")
            stats['found'] += 1
            results.append({
                'student_id': student_id,
                'self_grade': self_grade,
                'base_grade': base_grade,
                'source': 'submission_info.xlsx'
            })
        else:
            print(f"[NONE] Student {student_id}: No self-grade in Excel, using base grade {base_grade}")
            self_grade = base_grade
            stats['missing'] += 1
            stats['using_base'] += 1
            results.append({
                'student_id': student_id,
                'self_grade': self_grade,
                'base_grade': base_grade,
                'source': 'Base Grade (no Excel self-grade)'
            })
    else:
        print(f"[NONE] Student {student_id}: No submission_info.xlsx found, using base grade {base_grade}")
        self_grade = base_grade
        stats['missing'] += 1
        stats['using_base'] += 1
        results.append({
            'student_id': student_id,
            'self_grade': self_grade,
            'base_grade': base_grade,
            'source': 'Base Grade (no Excel file)'
        })

    # Update Excel
    ws.cell(row_num, self_grade_col).value = self_grade
    stats['processed'] += 1

# Save workbook
wb.save(excel_path)

# Print summary
print("\n" + "=" * 60)
print("SUMMARY REPORT")
print("=" * 60)
print(f"\nStudents Processed: {stats['processed']}")
print(f"Self-Grades Found in Excel: {stats['found']} ({stats['found']/stats['processed']*100:.1f}%)")
print(f"Self-Grades Missing: {stats['missing']} ({stats['missing']/stats['processed']*100:.1f}%)")
print(f"Using Base Grade as Default: {stats['using_base']}")

# Show students with self-grades different from base
print(f"\n[DETAILS] Students with self-grades:")
for result in results:
    if result['source'] == 'submission_info.xlsx':
        diff = result['self_grade'] - result['base_grade']
        if diff > 0:
            status = f"Overconfident by {diff}"
        elif diff < 0:
            status = f"Humble by {abs(diff)}"
        else:
            status = "Accurate!"
        print(f"  Student {result['student_id']}: Self={result['self_grade']}, Base={result['base_grade']} - {status}")

print(f"\n[OK] Excel updated: {excel_path}")
print(f"  Self-Grade column populated for {stats['processed']} students")
print("\n" + "=" * 60)
