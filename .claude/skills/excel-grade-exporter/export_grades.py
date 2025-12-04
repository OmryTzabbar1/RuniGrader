"""
Excel Grade Exporter
Generate Moodle-compatible Excel files with weighted grades
"""
import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from weighted_calculator import get_weighted_grade_for_student


def load_assessment_data(assignment_number: int) -> dict:
    """Load all assessment JSON files for an assignment."""
    assessment_dir = Path(f"assessments_tier2_assignment{assignment_number}")

    if not assessment_dir.exists():
        raise FileNotFoundError(f"Assessment directory not found: {assessment_dir}")

    assessments = {}
    json_files = list(assessment_dir.glob("tier2_assessment_*.json"))

    print(f"Loading {len(json_files)} assessment files...")

    for json_file in json_files:
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                student_id = data.get('student_id')
                if student_id:
                    assessments[student_id] = data
        except Exception as e:
            print(f"Error loading {json_file}: {e}")

    print(f"[OK] Loaded {len(assessments)} assessments")
    return assessments


def generate_feedback(assessment: dict) -> str:
    """Generate concise feedback from assessment data."""
    weaknesses = assessment.get('overall_assessment', {}).get('weaknesses', [])
    recommendations = assessment.get('recommended_actions', {})

    feedback_parts = []

    # Add main weaknesses
    if weaknesses and weaknesses[0] != "No critical weaknesses (<5/10)":
        # Get top 3 weaknesses
        for weakness in weaknesses[:3]:
            feedback_parts.append(weakness)

    # Add immediate actions if critical
    immediate = recommendations.get('immediate', [])
    if immediate and immediate[0] != "All critical areas addressed - focus on optimization":
        feedback_parts.extend(immediate[:2])

    if not feedback_parts:
        return "Good work! Continue improving based on recommendations."

    return "\n".join(feedback_parts)


def create_moodle_excel(assessments: dict, weighted_grades: dict, assignment_number: int):
    """Create Excel file in Moodle upload format."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Assignment {assignment_number} Grades"

    # Hebrew column headers (matching work1-runi-grade-format.csv)
    headers = [
        "מספר מזהה",              # Student ID
        "מצב",                    # Status
        "ציונים",                 # Grade
        "ציון מירבי",             # Max Grade
        "הציון ניתן לשינוי",      # Editable
        "מועד הגשת מטלה (אחרון)", # Submission Date
        "תאריך שינוי אחרון (ציון)", # Last Modified
        "הערות למשוב"            # Feedback
    ]

    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add student data
    for student_id in sorted(assessments.keys()):
        assessment = assessments[student_id]
        weighted_data = weighted_grades.get(student_id, {})

        weighted_grade = weighted_data.get('weighted_grade', assessment['total_score'])
        feedback = generate_feedback(assessment)
        assessment_date = assessment.get('assessment_date', '')

        row = [
            f"משתתף:{student_id}",                    # Student ID
            "הוגש למתן ציון -  - ",                  # Status
            round(weighted_grade, 2),                # Weighted Grade
            100,                                      # Max Grade
            "כן",                                     # Editable
            assessment_date,                          # Submission Date
            assessment_date,                          # Last Modified
            feedback                                  # Feedback
        ]

        ws.append(row)

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 60

    # Center align grade columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Wrap text in feedback column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Save to StudentGradesMoodleFormat directory
    output_dir = Path("StudentGradesMoodleFormat")
    output_dir.mkdir(exist_ok=True)

    filename = output_dir / f"Assignment{assignment_number}_Moodle_Grades.xlsx"
    wb.save(filename)
    print(f"[OK] Created: {filename}")
    return str(filename)


def create_comparison_excel(assessments: dict, weighted_grades: dict, assignment_number: int):
    """Create comparison Excel showing calculated vs weighted grades."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade Comparison"

    # Headers
    headers = [
        "Student ID",
        "Calculated Grade\n(10 Skills)",
        "Weighted Grade\n(Final)",
        "Difference",
        "Self-Grade",
        "Penalty",
        "Performance Tier"
    ]

    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add student data
    for student_id in sorted(assessments.keys()):
        assessment = assessments[student_id]
        weighted_data = weighted_grades.get(student_id, {})

        calculated_grade = assessment['total_score']
        weighted_grade = weighted_data.get('weighted_grade', calculated_grade)
        self_grade = weighted_data.get('self_grade')
        penalty = weighted_data.get('penalty', 0)
        difference = weighted_grade - calculated_grade
        tier = assessment['performance_tier']

        row = [
            student_id,
            round(calculated_grade, 2),
            round(weighted_grade, 2),
            round(difference, 2),
            round(self_grade, 2) if self_grade else "N/A",
            round(penalty, 2),
            tier
        ]

        ws.append(row)

        row_num = ws.max_row

        # Color-code calculated grade (column B)
        calculated_cell = ws.cell(row=row_num, column=2)

        if calculated_grade >= 90:
            calculated_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif calculated_grade >= 80:
            calculated_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        elif calculated_grade >= 70:
            calculated_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif calculated_grade >= 60:
            calculated_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            calculated_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

        # Color-code weighted grade (column C)
        weighted_cell = ws.cell(row=row_num, column=3)

        if weighted_grade >= 90:
            weighted_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif weighted_grade >= 80:
            weighted_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        elif weighted_grade >= 70:
            weighted_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif weighted_grade >= 60:
            weighted_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            weighted_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

        # Color-code difference
        diff_cell = ws.cell(row=row_num, column=4)
        if difference < 0:
            diff_cell.font = Font(color="FF0000")  # Red for negative
        elif difference > 0:
            diff_cell.font = Font(color="00AA00")  # Green for positive

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18

    # Center align all data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add summary statistics at the bottom
    ws.append([])
    ws.append(["Summary Statistics"])

    calculated_avg = sum(a['total_score'] for a in assessments.values()) / len(assessments)
    weighted_avg = sum(w.get('weighted_grade', 0) for w in weighted_grades.values()) / len(weighted_grades)
    penalty_avg = sum(w.get('penalty', 0) for w in weighted_grades.values()) / len(weighted_grades)

    ws.append(["Average Calculated Grade:", round(calculated_avg, 2)])
    ws.append(["Average Weighted Grade:", round(weighted_avg, 2)])
    ws.append(["Average Penalty:", round(penalty_avg, 2)])

    # Save to StudentGradesForRami directory
    output_dir = Path("StudentGradesForRami")
    output_dir.mkdir(exist_ok=True)

    filename = output_dir / f"Assignment{assignment_number}_Grade_Comparison.xlsx"
    wb.save(filename)
    print(f"[OK] Created: {filename}")
    return str(filename)


def main():
    if len(sys.argv) < 2:
        print("Usage: python export_grades.py <assignment_number>")
        print("Example: python export_grades.py 1")
        sys.exit(1)

    assignment_number = int(sys.argv[1])
    submissions_dir = f"WorkSubmissions0{assignment_number}"

    print(f"\n{'='*60}")
    print(f"  Excel Grade Exporter - Assignment {assignment_number}")
    print(f"{'='*60}\n")

    # Load assessment data
    assessments = load_assessment_data(assignment_number)

    if not assessments:
        print("❌ No assessments found!")
        sys.exit(1)

    # Calculate weighted grades
    print(f"\nCalculating weighted grades using submissions from {submissions_dir}...")
    weighted_grades = {}

    for student_id, assessment in assessments.items():
        base_grade = assessment['total_score']
        weighted_grade, penalty, self_grade = get_weighted_grade_for_student(
            student_id, base_grade, submissions_dir
        )

        weighted_grades[student_id] = {
            'weighted_grade': weighted_grade,
            'penalty': penalty,
            'self_grade': self_grade
        }

        status = "[OK]" if self_grade is not None else "[WARN]"
        print(f"  {status} Student {student_id}: Base={base_grade:.1f}, Weighted={weighted_grade:.1f}, Penalty={penalty:.1f}")

    print(f"\n[OK] Calculated weighted grades for {len(weighted_grades)} students")

    # Generate Excel files
    print("\nGenerating Excel files...")
    moodle_file = create_moodle_excel(assessments, weighted_grades, assignment_number)
    comparison_file = create_comparison_excel(assessments, weighted_grades, assignment_number)

    # Summary
    print(f"\n{'='*60}")
    print("  Summary")
    print(f"{'='*60}")
    print(f"Students Processed: {len(assessments)}")
    print(f"Files Created:")
    print(f"  1. {moodle_file} (Moodle upload format)")
    print(f"  2. {comparison_file} (Grade comparison)")
    print(f"\n[SUCCESS] Export complete!")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
