#!/usr/bin/env python3
"""
Weighted Grade Calculator - Apply exponential self-grading penalty formula

Extracts self-proclaimed grades from student PDFs and calculates weighted final grades
using the exponential penalty formula that rewards accurate/humble self-assessment.
"""

import os
import sys
import math
import re
import argparse
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from PyPDF2 import PdfReader
except ImportError:
    print("Warning: PyPDF2 not installed. Will skip PDF text extraction. Run: pip install PyPDF2")
    PdfReader = None


# Constants from grading formula
SCALE_COEFFICIENT_A = 0.086603
SCALE_EXPONENT_B = 0.027465
TOTAL_REQUIREMENTS = 22
MIN_SELF_GRADE = 60
MAX_SELF_GRADE = 100


class WeightedGradeCalculator:
    """Calculate weighted grades using exponential penalty formula."""

    def __init__(self):
        self.stats = {
            'processed': 0,
            'self_grades_found': 0,
            'self_grades_missing': 0,
            'no_penalty': 0,
            'small_penalty': 0,
            'medium_penalty': 0,
            'large_penalty': 0,
            'penalties': []
        }

    def calculate_scale(self, self_grade: int) -> float:
        """Calculate exponential scale multiplier."""
        return SCALE_COEFFICIENT_A * math.exp(SCALE_EXPONENT_B * self_grade)

    def calculate_weighted_grade(self, self_grade: int, base_grade: float) -> dict:
        """
        Calculate weighted grade with penalty for overconfidence.

        Args:
            self_grade: Student's self-proclaimed grade (60-100)
            base_grade: Actual grade from Tier 2 assessment (0-100)

        Returns:
            dict with: weighted_grade, penalty, scale, difference
        """
        # Validate inputs
        if not (MIN_SELF_GRADE <= self_grade <= MAX_SELF_GRADE):
            raise ValueError(f"Self-grade {self_grade} out of valid range [{MIN_SELF_GRADE}, {MAX_SELF_GRADE}]")

        if not (0 <= base_grade <= 100):
            raise ValueError(f"Base grade {base_grade} out of valid range [0, 100]")

        # Calculate scale
        scale = self.calculate_scale(self_grade)

        # Calculate difference
        difference = self_grade - base_grade

        # Apply penalty only if overestimated
        if difference > 0:
            penalty = difference * scale
            weighted_grade = max(0, base_grade - penalty)
        else:
            # Reward humility - no penalty
            penalty = 0
            weighted_grade = base_grade

        return {
            'weighted_grade': weighted_grade,
            'penalty': penalty,
            'scale': scale,
            'difference': difference
        }

    def extract_self_grade_from_pdf(self, pdf_path: str) -> int:
        """
        Extract self-proclaimed grade from student submission PDF.

        Searches for patterns like:
        - "Self-Grade: 85"
        - "I estimate: 90%"
        - "My grade: 80"
        """
        if not PdfReader or not os.path.exists(pdf_path):
            return None

        try:
            reader = PdfReader(pdf_path)

            # Try extracting all text
            text = ""
            for page in reader.pages:
                text += page.extract_text()

            # Search patterns
            patterns = [
                r"Self[- ]?Grade:?\s*(\d+)",
                r"I estimate:?\s*(\d+)%?",
                r"My grade:?\s*(\d+)",
                r"Self[- ]?assessment:?\s*(\d+)",
                r"Expected grade:?\s*(\d+)",
                r"Predicted grade:?\s*(\d+)"
            ]

            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    grade = int(match.group(1))
                    # Validate range
                    if MIN_SELF_GRADE <= grade <= MAX_SELF_GRADE:
                        return grade

            return None

        except Exception as e:
            print(f"Warning: Could not extract from {pdf_path}: {e}")
            return None

    def find_student_pdfs(self, submissions_dir: str) -> dict:
        """Find all student submission PDFs."""
        pdf_map = {}

        submissions_path = Path(submissions_dir)
        if not submissions_path.exists():
            print(f"Error: Submissions directory not found: {submissions_dir}")
            return pdf_map

        for participant_dir in submissions_path.glob("Participant_*"):
            if not participant_dir.is_dir():
                continue

            # Extract student ID
            match = re.search(r"Participant_(\d+)", participant_dir.name)
            if not match:
                continue

            student_id = match.group(1)

            # Find submission PDF (not grade report)
            for pdf_file in participant_dir.glob("*.pdf"):
                # Skip generated grade reports
                if "Student_Grade_Report" in pdf_file.name:
                    continue

                pdf_map[student_id] = str(pdf_file)
                break

        return pdf_map

    def process_excel(self, excel_path: str, submissions_dir: str, output_path: str = None):
        """
        Process Excel file: extract self-grades, calculate weighted grades, update spreadsheet.
        """
        if not os.path.exists(excel_path):
            print(f"Error: Excel file not found: {excel_path}")
            return False

        # Backup original file
        backup_path = excel_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        shutil.copy2(excel_path, backup_path)
        print(f"✓ Backup created: {backup_path}")

        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Find student PDF files
        print(f"\n🔍 Scanning for student submission PDFs in {submissions_dir}...")
        pdf_map = self.find_student_pdfs(submissions_dir)
        print(f"✓ Found {len(pdf_map)} student PDFs")

        # Find column indices
        header_row = 1
        headers = {cell.value: cell.column for cell in ws[header_row] if cell.value}

        student_id_col = headers.get('Student ID')
        generated_grade_col = headers.get('Generated\nGrade (/100)')
        weighted_grade_col = headers.get('Weighted\nGrade')

        if not all([student_id_col, generated_grade_col, weighted_grade_col]):
            print(f"Error: Required columns not found in Excel")
            print(f"Found headers: {list(headers.keys())}")
            return False

        # Insert Self-Grade column after Student ID (if not exists)
        self_grade_col = headers.get('Self-Grade')
        if not self_grade_col:
            # Insert new column B
            ws.insert_cols(student_id_col + 1)
            self_grade_col = student_id_col + 1
            ws.cell(header_row, self_grade_col).value = "Self-Grade"
            ws.cell(header_row, self_grade_col).font = Font(bold=True, color="FFFFFF", size=11)
            ws.cell(header_row, self_grade_col).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.cell(header_row, self_grade_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Adjust other column references
            generated_grade_col += 1
            weighted_grade_col += 1

        # Insert Penalty column before Weighted Grade (if not exists)
        penalty_col = headers.get('Penalty')
        if not penalty_col:
            ws.insert_cols(weighted_grade_col)
            penalty_col = weighted_grade_col
            weighted_grade_col += 1

            ws.cell(header_row, penalty_col).value = "Penalty"
            ws.cell(header_row, penalty_col).font = Font(bold=True, color="FFFFFF", size=11)
            ws.cell(header_row, penalty_col).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.cell(header_row, penalty_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        print(f"\n📊 Processing students...")

        # Process each student row
        for row_num in range(header_row + 1, ws.max_row + 1):
            student_id = ws.cell(row_num, student_id_col).value
            if not student_id:
                continue

            student_id = str(student_id)
            generated_grade = ws.cell(row_num, generated_grade_col).value

            if not generated_grade:
                print(f"⚠️  Student {student_id}: No generated grade found")
                continue

            base_grade = float(generated_grade)

            # Extract self-grade
            self_grade = None

            # Try PDF extraction
            if student_id in pdf_map:
                self_grade = self.extract_self_grade_from_pdf(pdf_map[student_id])
                if self_grade:
                    print(f"✓ Student {student_id}: Self-grade {self_grade} extracted from PDF")
                    self.stats['self_grades_found'] += 1

            # If not found, use base grade as default (no penalty)
            if self_grade is None:
                self_grade = int(base_grade)
                print(f"⚠️  Student {student_id}: No self-grade found, using base grade {self_grade} (no penalty)")
                self.stats['self_grades_missing'] += 1

            # Calculate weighted grade
            result = self.calculate_weighted_grade(self_grade, base_grade)

            # Track penalty statistics
            penalty = result['penalty']
            if penalty == 0:
                self.stats['no_penalty'] += 1
            elif penalty <= 5:
                self.stats['small_penalty'] += 1
            elif penalty <= 15:
                self.stats['medium_penalty'] += 1
            else:
                self.stats['large_penalty'] += 1

            self.stats['penalties'].append({
                'student_id': student_id,
                'self_grade': self_grade,
                'base_grade': base_grade,
                'penalty': penalty,
                'weighted_grade': result['weighted_grade']
            })

            # Update Excel
            ws.cell(row_num, self_grade_col).value = self_grade
            ws.cell(row_num, self_grade_col).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row_num, penalty_col).value = round(-penalty, 2) if penalty > 0 else 0
            ws.cell(row_num, penalty_col).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row_num, weighted_grade_col).value = round(result['weighted_grade'], 2)
            ws.cell(row_num, weighted_grade_col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row_num, weighted_grade_col).font = Font(bold=True, size=11)

            # Color code weighted grade
            if result['weighted_grade'] >= 90:
                ws.cell(row_num, weighted_grade_col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif result['weighted_grade'] >= 80:
                ws.cell(row_num, weighted_grade_col).fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif result['weighted_grade'] >= 70:
                ws.cell(row_num, weighted_grade_col).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif result['weighted_grade'] >= 60:
                ws.cell(row_num, weighted_grade_col).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                ws.cell(row_num, weighted_grade_col).fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

            self.stats['processed'] += 1

        # Save workbook
        output_file = output_path if output_path else excel_path
        wb.save(output_file)
        print(f"\n✓ Excel updated: {output_file}")

        return True

    def print_summary(self):
        """Print summary statistics."""
        print("\n" + "="*60)
        print("╔════════════════════════════════════════════════╗")
        print("║  WEIGHTED GRADE CALCULATION SUMMARY           ║")
        print("╚════════════════════════════════════════════════╝")
        print()

        print(f"Students Processed: {self.stats['processed']}")
        print(f"Self-Grades Found: {self.stats['self_grades_found']} ({self.stats['self_grades_found']/self.stats['processed']*100:.1f}%)")
        print(f"Self-Grades Missing: {self.stats['self_grades_missing']} ({self.stats['self_grades_missing']/self.stats['processed']*100:.1f}%)")
        print()

        print("Penalty Statistics:")
        print(f"├─ No Penalty (accurate/humble): {self.stats['no_penalty']} ({self.stats['no_penalty']/self.stats['processed']*100:.1f}%)")
        print(f"├─ Small Penalty (1-5 points): {self.stats['small_penalty']} ({self.stats['small_penalty']/self.stats['processed']*100:.1f}%)")
        print(f"├─ Medium Penalty (6-15 points): {self.stats['medium_penalty']} ({self.stats['medium_penalty']/self.stats['processed']*100:.1f}%)")
        print(f"└─ Large Penalty (>15 points): {self.stats['large_penalty']} ({self.stats['large_penalty']/self.stats['processed']*100:.1f}%)")
        print()

        # Calculate averages
        penalties = self.stats['penalties']
        if penalties:
            avg_base = sum(p['base_grade'] for p in penalties) / len(penalties)
            avg_weighted = sum(p['weighted_grade'] for p in penalties) / len(penalties)
            avg_penalty = sum(p['penalty'] for p in penalties) / len(penalties)

            print(f"Average Base Grade: {avg_base:.2f}")
            print(f"Average Weighted Grade: {avg_weighted:.2f}")
            print(f"Average Penalty: {avg_penalty:.2f} points")
            print()

            # Top self-assessors (most accurate)
            sorted_by_accuracy = sorted(penalties, key=lambda p: abs(p['self_grade'] - p['base_grade']))
            print("Top 3 Self-Assessors (most accurate):")
            for i, p in enumerate(sorted_by_accuracy[:3], 1):
                diff = abs(p['self_grade'] - p['base_grade'])
                print(f"{i}. Student {p['student_id']}: Self={p['self_grade']}, Actual={p['base_grade']:.0f}, Diff={diff:.1f} ✓")
            print()

            # Bottom self-assessors (most overconfident)
            overconfident = [p for p in penalties if p['penalty'] > 0]
            if overconfident:
                sorted_by_overconfidence = sorted(overconfident, key=lambda p: p['penalty'], reverse=True)
                print("Top 3 Most Overconfident:")
                for i, p in enumerate(sorted_by_overconfidence[:3], 1):
                    print(f"{i}. Student {p['student_id']}: Self={p['self_grade']}, Actual={p['base_grade']:.0f}, Penalty={p['penalty']:.1f} ❌")

        print("\n" + "="*60)


def main():
    parser = argparse.ArgumentParser(
        description='Calculate weighted grades using exponential self-grading penalty formula',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process Excel with student submissions
  python calculate_weighted_grades.py \\
    --excel Assignment3_Grading_Summary.xlsx \\
    --submissions WorkSubmissions01

  # Process and save to new file
  python calculate_weighted_grades.py \\
    --excel Assignment3_Grading_Summary.xlsx \\
    --submissions WorkSubmissions01 \\
    --output Assignment3_Grading_Summary_Weighted.xlsx
        """
    )

    parser.add_argument('--excel', required=True, help='Path to Excel grading summary')
    parser.add_argument('--submissions', required=True, help='Path to student submissions directory')
    parser.add_argument('--output', help='Output Excel file (default: overwrite input)')

    args = parser.parse_args()

    # Create calculator
    calculator = WeightedGradeCalculator()

    # Process Excel
    success = calculator.process_excel(args.excel, args.submissions, args.output)

    if success:
        # Print summary
        calculator.print_summary()
        print("\n✅ Weighted grade calculation complete!")
        return 0
    else:
        print("\n❌ Weighted grade calculation failed")
        return 1


if __name__ == '__main__':
    sys.exit(main())
