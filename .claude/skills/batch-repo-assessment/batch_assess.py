#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Batch assess all student GitHub repositories
Extracts GitHub URLs from Excel files and runs git-repo-assessment for each
"""

import os
import sys
import json
from pathlib import Path
import openpyxl

# Set UTF-8 encoding for stdout
if sys.platform == 'win32':
    import codecs
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    if sys.stderr.encoding != 'utf-8':
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')


def extract_github_urls_from_excel(submissions_folder):
    """Extract GitHub URLs and participant info from all Excel files."""
    submissions_folder = Path(submissions_folder)

    results = []

    # Find all submission_info.xlsx files
    excel_files = list(submissions_folder.glob("*/submission_info.xlsx"))

    print(f"Found {len(excel_files)} Excel files to process\n")

    for excel_path in sorted(excel_files):
        folder_path = excel_path.parent
        participant_id = folder_path.name.split('_')[1]

        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            # Extract data from Excel (row format: Field | Value)
            # Row 2: Participant ID
            # Row 3: Group Code
            # Row 6: GitHub Repository (row 7 in 1-indexed)
            # Row 7: Suggested Grade

            github_url = ws.cell(6, 2).value  # GitHub Repository is in row 6, column 2
            group_code = ws.cell(3, 2).value  # Group Code is in row 3, column 2

            wb.close()

            if github_url and github_url.startswith('http'):
                results.append({
                    'participant_id': participant_id,
                    'folder_path': str(folder_path),
                    'github_url': github_url,
                    'group_code': group_code
                })
                print(f"[OK] {participant_id}: {github_url}")
            else:
                print(f"[SKIP] {participant_id}: No valid GitHub URL found")

        except Exception as e:
            print(f"[ERROR] {participant_id}: {e}")

    return results


def save_assessment_batch_file(results, output_file='run_assessments.json'):
    """Save the batch assessment data to JSON for processing."""
    output_path = Path(output_file)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print(f"\n[OK] Saved {len(results)} repositories to {output_path}")
    print(f"\nReady to assess {len(results)} repositories")

    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python batch_assess.py <submissions_folder> [output_json]")
        print("Example: python batch_assess.py WorkSubmissions01")
        sys.exit(1)

    submissions_folder = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'run_assessments.json'

    # Extract all GitHub URLs
    results = extract_github_urls_from_excel(submissions_folder)

    # Save to JSON
    if results:
        save_assessment_batch_file(results, output_file)
    else:
        print("[ERROR] No valid GitHub URLs found")
        sys.exit(1)
