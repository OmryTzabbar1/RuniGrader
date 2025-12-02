#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Grade Extractor - Processes student submission PDFs and creates Excel spreadsheet
This script is designed to work with Claude Code agent
"""

import os
import sys
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

# Set UTF-8 encoding for stdout
if sys.platform == 'win32':
    import codecs
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    if sys.stderr.encoding != 'utf-8':
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')


def create_excel_workbook(output_file):
    """Create a new Excel workbook with formatted headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Submissions"

    # Define headers
    headers = [
        "Participant ID",
        "Group Code",
        "Student 1",
        "Student 2",
        "GitHub Repository",
        "Suggested Grade",
        "PDF Filename"
    ]

    # Write headers with formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Set column widths
    ws.column_dimensions['A'].width = 15  # Participant ID
    ws.column_dimensions['B'].width = 20  # Group Code
    ws.column_dimensions['C'].width = 35  # Student 1
    ws.column_dimensions['D'].width = 35  # Student 2
    ws.column_dimensions['E'].width = 60  # GitHub
    ws.column_dimensions['F'].width = 15  # Grade
    ws.column_dimensions['G'].width = 40  # Filename

    # Freeze the header row
    ws.freeze_panes = 'A2'

    return wb


def add_submission_to_excel(ws, row, participant_id, group_code, student1, student2, github, grade, pdf_filename):
    """Add a submission row to the Excel worksheet with formatting."""
    # Add data
    ws.cell(row=row, column=1, value=participant_id)
    ws.cell(row=row, column=2, value=group_code)
    ws.cell(row=row, column=3, value=student1)
    ws.cell(row=row, column=4, value=student2)
    ws.cell(row=row, column=5, value=github)
    ws.cell(row=row, column=6, value=grade)
    ws.cell(row=row, column=7, value=pdf_filename)

    # Apply alignment
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Make GitHub link clickable if it's a valid URL
    if github and github.startswith('http'):
        github_cell = ws.cell(row=row, column=5)
        github_cell.hyperlink = github
        github_cell.font = Font(color="0563C1", underline="single")


def get_submission_folders(base_dir):
    """Get all submission folders from the base directory."""
    base_path = Path(base_dir)
    folders = []

    if base_path.is_dir():
        # Single submission folder
        submission_folders = sorted([
            d for d in base_path.iterdir()
            if d.is_dir() and d.name.startswith('Participant_')
        ])
        folders.extend([(base_path.name, folder) for folder in submission_folders])
    else:
        print(f"Error: {base_dir} is not a directory")

    return folders


def extract_participant_id(folder_name):
    """Extract participant ID from folder name like 'Participant_38950_assignsubmission_file'"""
    match = re.search(r'Participant_(\d+)', folder_name)
    return match.group(1) if match else folder_name


def find_pdf_in_folder(folder_path):
    """Find the first PDF file in the given folder."""
    pdf_files = list(folder_path.glob('*.pdf'))
    return pdf_files[0] if pdf_files else None


def save_submission_list(submissions, output_file):
    """Save the list of submissions to a JSON file for the agent to process."""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(submissions, f, indent=2, ensure_ascii=False)


def main():
    """Main function to orchestrate the extraction process."""
    if len(sys.argv) < 3:
        print("Usage: python extract_grades.py <submissions_folder> <output_excel_file>")
        print("\nExample:")
        print("  python extract_grades.py WorkSubmissions01 grades.xlsx")
        sys.exit(1)

    base_dir = sys.argv[1]
    output_excel = sys.argv[2]

    # Check if base directory exists
    if not os.path.exists(base_dir):
        print(f"Error: Directory '{base_dir}' does not exist")
        sys.exit(1)

    # Create Excel workbook
    print(f"Creating Excel workbook: {output_excel}")
    wb = create_excel_workbook(output_excel)
    ws = wb.active

    # Scan for submission folders
    print(f"\nScanning for submissions in: {base_dir}")
    base_path = Path(base_dir)
    submission_folders = sorted([
        d for d in base_path.iterdir()
        if d.is_dir() and d.name.startswith('Participant_')
    ])

    print(f"Found {len(submission_folders)} submission folders\n")

    # Create submissions list for agent processing
    submissions = []

    for idx, folder in enumerate(submission_folders, start=1):
        participant_id = extract_participant_id(folder.name)
        pdf_file = find_pdf_in_folder(folder)

        if not pdf_file:
            print(f"WARNING: {idx}. {folder.name} - No PDF found")
            continue

        submission_info = {
            'index': idx,
            'participant_id': participant_id,
            'folder_name': folder.name,
            'pdf_path': str(pdf_file.absolute()),
            'pdf_filename': pdf_file.name
        }
        submissions.append(submission_info)
        print(f"[{idx}] Participant {participant_id} - {pdf_file.name}")

    # Save submission list
    submissions_json = output_excel.replace('.xlsx', '_submissions.json')
    save_submission_list(submissions, submissions_json)

    # Save initial Excel with just participant IDs and filenames
    current_row = 2
    for submission in submissions:
        add_submission_to_excel(
            ws, current_row,
            participant_id=submission['participant_id'],
            group_code='',
            student1='',
            student2='',
            github='',
            grade='',
            pdf_filename=submission['pdf_filename']
        )
        current_row += 1

    wb.save(output_excel)

    print(f"\n[OK] Initial Excel file created: {output_excel}")
    print(f"[OK] Submission list saved: {submissions_json}")
    print(f"\nTotal submissions found: {len(submissions)}")
    print("\n" + "="*60)
    print("Next step: Run the agent to extract information from PDFs")
    print("="*60)

    # Return the submissions list for the agent to process
    return submissions


if __name__ == '__main__':
    main()
