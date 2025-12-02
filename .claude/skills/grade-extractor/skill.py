"""
Grade Extractor Skill for Claude CLI
Extracts student submission information from PDFs and populates an Excel file.
"""

import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re


def extract_info_from_pdf_text(pdf_text):
    """
    Extract the 5 required pieces of information from PDF text.
    Returns a dict with: group_code, student1, student2, github_link, suggested_grade
    """
    info = {
        'group_code': '',
        'student1': '',
        'student2': '',
        'github_link': '',
        'suggested_grade': ''
    }

    # Pattern 1: Direct numbered format (1. Group code:, 2. Student one:, etc.)
    patterns = {
        'group_code': [
            r'1\.\s*(?:Group [Cc]ode|Codegroupe|קוד קבוצה):\s*([^\n]+)',
            r'Group [Cc]ode:\s*([^\n]+)',
        ],
        'student1': [
            r'2\.\s*(?:Student one|Member A):\s*([^\n]+)',
            r'Student one:\s*([^\n]+)',
        ],
        'student2': [
            r'3\.\s*(?:Student two|Member B):\s*([^\n]+)',
            r'Student two:\s*([^\n]+)',
        ],
        'github_link': [
            r'4\.\s*(?:Repo link|GitHub Repository):\s*([^\n]+)',
            r'GitHub Repository:\s*([^\n]+)',
            r'Repo link\s+([^\n]+)',
        ],
        'suggested_grade': [
            r'5\.\s*(?:Grade suggestion|הציון העצמי שלי):\s*(\d+)',
            r'Grade suggestion:\s*(\d+)',
            r'הציון העצמי שלי\s*(\d+)/100',
        ]
    }

    # Try each pattern for each field
    for field, pattern_list in patterns.items():
        for pattern in pattern_list:
            match = re.search(pattern, pdf_text, re.IGNORECASE)
            if match:
                info[field] = match.group(1).strip()
                break

    # Clean up the extracted data
    info['group_code'] = info['group_code'].strip()
    info['student1'] = info['student1'].strip()
    info['student2'] = info['student2'].strip()

    # Extract GitHub URL if it's embedded
    github_match = re.search(r'https://github\.com/[^\s\)]+', info['github_link'])
    if github_match:
        info['github_link'] = github_match.group(0)

    # Extract just the number from suggested grade
    grade_match = re.search(r'(\d+)', info['suggested_grade'])
    if grade_match:
        info['suggested_grade'] = grade_match.group(1)

    return info


def create_excel_workbook(output_file):
    """Create a new Excel workbook with headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Submissions"

    # Define headers
    headers = [
        "Participant ID",
        "Group Code",
        "Student 1",
        "Student 2",
        "GitHub Repository",
        "Suggested Grade",
        "PDF Filename"
    ]

    # Write headers with formatting
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 15  # Participant ID
    ws.column_dimensions['B'].width = 20  # Group Code
    ws.column_dimensions['C'].width = 30  # Student 1
    ws.column_dimensions['D'].width = 30  # Student 2
    ws.column_dimensions['E'].width = 50  # GitHub
    ws.column_dimensions['F'].width = 15  # Grade
    ws.column_dimensions['G'].width = 40  # Filename

    wb.save(output_file)
    return wb


def process_submissions(base_dir, output_file):
    """
    Process all student submissions in the base directory.
    Each submission folder contains a PDF with student information.
    """
    base_path = Path(base_dir)

    # Create or load workbook
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        current_row = ws.max_row + 1
    else:
        wb = create_excel_workbook(output_file)
        ws = wb.active
        current_row = 2

    # Find all submission folders
    submission_folders = sorted([d for d in base_path.iterdir() if d.is_dir() and d.name.startswith('Participant_')])

    print(f"Found {len(submission_folders)} submission folders")

    for folder in submission_folders:
        # Extract participant ID from folder name
        participant_id = folder.name.split('_')[1]

        # Find PDF file in folder
        pdf_files = list(folder.glob('*.pdf'))

        if not pdf_files:
            print(f"Warning: No PDF found in {folder.name}")
            continue

        pdf_file = pdf_files[0]
        print(f"Processing: {folder.name} - {pdf_file.name}")

        # Read PDF using Claude Code's Read tool
        # For this skill, we'll indicate that the PDF needs to be read
        info = {
            'participant_id': participant_id,
            'pdf_path': str(pdf_file),
            'pdf_filename': pdf_file.name
        }

        # Write to Excel (leaving extraction fields empty for now - will be filled by agent)
        ws.cell(row=current_row, column=1, value=participant_id)
        ws.cell(row=current_row, column=2, value='')  # Group Code
        ws.cell(row=current_row, column=3, value='')  # Student 1
        ws.cell(row=current_row, column=4, value='')  # Student 2
        ws.cell(row=current_row, column=5, value='')  # GitHub
        ws.cell(row=current_row, column=6, value='')  # Grade
        ws.cell(row=current_row, column=7, value=pdf_file.name)

        current_row += 1

    wb.save(output_file)
    print(f"\nExcel file created: {output_file}")
    print(f"Total submissions processed: {len(submission_folders)}")

    return len(submission_folders)


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python skill.py <submissions_folder> <output_excel_file>")
        sys.exit(1)

    base_dir = sys.argv[1]
    output_file = sys.argv[2]

    if not os.path.exists(base_dir):
        print(f"Error: Directory {base_dir} does not exist")
        sys.exit(1)

    process_submissions(base_dir, output_file)
