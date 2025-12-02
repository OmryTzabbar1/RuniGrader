#!/usr/bin/env python3
"""
Update Excel with extracted submission data
Usage: python update_excel.py <excel_file> <participant_id> <group_code> <student1> <student2> <github> <grade>
"""

import sys
import openpyxl
from openpyxl.styles import Font, Alignment


def find_participant_row(ws, participant_id):
    """Find the row number for a given participant ID."""
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and str(cell_value) == str(participant_id):
            return row
    return None


def update_submission_data(excel_file, participant_id, group_code, student1, student2, github, grade):
    """Update a single submission row in the Excel file."""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        row = find_participant_row(ws, participant_id)
        if not row:
            print(f"Error: Participant {participant_id} not found in Excel file")
            return False

        # Update the cells
        ws.cell(row=row, column=2, value=group_code)  # Group Code
        ws.cell(row=row, column=3, value=student1)     # Student 1
        ws.cell(row=row, column=4, value=student2)     # Student 2
        ws.cell(row=row, column=5, value=github)       # GitHub
        ws.cell(row=row, column=6, value=grade)        # Grade

        # Format GitHub as hyperlink if it's a URL
        if github and github.startswith('http'):
            github_cell = ws.cell(row=row, column=5)
            github_cell.hyperlink = github
            github_cell.font = Font(color="0563C1", underline="single")

        # Apply text wrapping
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        wb.save(excel_file)
        print(f"[OK] Updated Participant {participant_id}")
        return True

    except Exception as e:
        print(f"[ERROR] Error updating Participant {participant_id}: {str(e)}")
        return False


def main():
    if len(sys.argv) != 8:
        print("Usage: python update_excel.py <excel_file> <participant_id> <group_code> <student1> <student2> <github> <grade>")
        print("\nExample:")
        print('  python update_excel.py grades.xlsx 38950 "roeiandguy" "Roei Bracha 208933325" "Guy Bilitski 2087332532" "https://github.com/..." "100"')
        sys.exit(1)

    excel_file = sys.argv[1]
    participant_id = sys.argv[2]
    group_code = sys.argv[3]
    student1 = sys.argv[4]
    student2 = sys.argv[5]
    github = sys.argv[6]
    grade = sys.argv[7]

    success = update_submission_data(excel_file, participant_id, group_code, student1, student2, github, grade)
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
