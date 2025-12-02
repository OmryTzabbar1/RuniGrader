#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Create individual Excel files in each student's submission folder
Each Excel file contains the extracted information from that student's PDF
"""

import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Set UTF-8 encoding for stdout
if sys.platform == 'win32':
    import codecs
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    if sys.stderr.encoding != 'utf-8':
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')


def create_student_excel(folder_path, participant_id, group_code, student1, student2, github, grade, pdf_filename):
    """Create an individual Excel file for a student's submission."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submission Info"

    # Define headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_alignment = Alignment(vertical="top", wrap_text=True)

    # Add data in a vertical format (label - value pairs)
    rows = [
        ("Field", "Value"),
        ("Participant ID", participant_id),
        ("Group Code", group_code),
        ("Student 1", student1),
        ("Student 2", student2),
        ("GitHub Repository", github),
        ("Suggested Grade", grade),
        ("PDF Filename", pdf_filename),
    ]

    for row_idx, (label, value) in enumerate(rows, start=1):
        # Label column
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        if row_idx == 1:  # Header row
            label_cell.font = header_font
            label_cell.fill = header_fill
            label_cell.alignment = header_alignment
        else:
            label_cell.font = Font(bold=True)
            label_cell.alignment = data_alignment

        # Value column
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:  # Header row
            value_cell.font = header_font
            value_cell.fill = header_fill
            value_cell.alignment = header_alignment
        else:
            value_cell.alignment = data_alignment

            # Make GitHub link clickable
            if label == "GitHub Repository" and value and value.startswith('http'):
                value_cell.hyperlink = value
                value_cell.font = Font(color="0563C1", underline="single")

    # Set column widths
    ws.column_dimensions['A'].width = 25  # Label column
    ws.column_dimensions['B'].width = 60  # Value column

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Save the Excel file in the student's folder
    output_file = folder_path / "submission_info.xlsx"
    wb.save(output_file)

    return output_file


if __name__ == '__main__':
    if len(sys.argv) != 9:
        print("Usage: python create_individual_excel.py <folder_path> <participant_id> <group_code> <student1> <student2> <github> <grade> <pdf_filename>")
        sys.exit(1)

    folder_path = Path(sys.argv[1])
    participant_id = sys.argv[2]
    group_code = sys.argv[3]
    student1 = sys.argv[4]
    student2 = sys.argv[5]
    github = sys.argv[6]
    grade = sys.argv[7]
    pdf_filename = sys.argv[8]

    output_file = create_student_excel(folder_path, participant_id, group_code, student1, student2, github, grade, pdf_filename)
    print(f"[OK] Created {output_file}")
