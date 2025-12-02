#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Assignment 3 Grades"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Grade color fills
grade_a = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
grade_b = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
grade_c = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
grade_d = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
grade_f = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Headers
headers = [
    "Student ID",
    "Teammate IDs",
    "Team Name",
    "Estimated Grade\n(from .md)",
    "TRUE/FALSE\nScore",
    "Generated\nGrade (/100)",
    "Weighted\nGrade",
    "Assessment\nDate",
    "Key Strengths",
    "Key Weaknesses"
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Set column widths
column_widths = [12, 15, 25, 12, 12, 12, 12, 12, 50, 50]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# All student data
all_students = [
    {"id": "38950", "teammates": "38949, 38972, 38973", "team": "The_Awesome_Foursome", "estimated": "86%", "true_false": "19/22 (86%)", "grade": 87, "strengths": "Complete PRD, 130 tests, 654 docstrings, research notebook with plots, cost analysis, 27 screenshots, good git", "weaknesses": "No pre-commit hooks, could improve test coverage"},
    {"id": "38951", "teammates": "38952", "team": "Agents2025-2026", "estimated": "68%", "true_false": "15/22 (68%)", "grade": 68, "strengths": "Good documentation (382 docstrings), FastAPI implementation, decent git (29 commits), 10 screenshots", "weaknesses": "NO research, NO cost analysis, minimal testing (8 tests), no CI/CD"},
    {"id": "38952", "teammates": "38951", "team": "Agents2025-2026", "estimated": "68%", "true_false": "15/22 (68%)", "grade": 68, "strengths": "Good documentation (382 docstrings), FastAPI implementation, decent git (29 commits), 10 screenshots", "weaknesses": "NO research, NO cost analysis, minimal testing (8 tests), no CI/CD"},
    {"id": "38953", "teammates": "38949, 38972, 38973", "team": "The_Awesome_Foursome", "estimated": "86%", "true_false": "19/22 (86%)", "grade": 87, "strengths": "Complete PRD, 130 tests, 654 docstrings, research notebook with plots, cost analysis, 27 screenshots, good git", "weaknesses": "No pre-commit hooks, could improve test coverage"},
    {"id": "38954", "teammates": "38959, 38972", "team": "LLM_Agents_Tom_Igor_Roie", "estimated": "91%", "true_false": "20/22 (91%)", "grade": 86, "strengths": "144 tests, 413 docstrings, 3062-word README, complete prompt book, budget management, 14 screenshots, 55 commits", "weaknesses": "No research component, no pre-commit hooks"},
    {"id": "38955", "teammates": "38974", "team": "The_Sucsess", "estimated": "64%", "true_false": "14/22 (64%)", "grade": 67, "strengths": "Good documentation (368 docstrings), decent testing (33 tests), good git (23 commits), 7 screenshots", "weaknesses": "NO research, NO cost analysis, no architecture doc, no CI/CD"},
    {"id": "38956", "teammates": "38974", "team": "The_Sucsess", "estimated": "DUPLICATE", "true_false": "14/22 (64%)", "grade": 67, "strengths": "DUPLICATE of 38955 - Good documentation, decent testing, good git", "weaknesses": "NO research, NO cost analysis, no architecture doc"},
    {"id": "38957", "teammates": "38974", "team": "The_Sucsess", "estimated": "DUPLICATE", "true_false": "14/22 (64%)", "grade": 67, "strengths": "DUPLICATE of 38955 - Good documentation, decent testing, good git", "weaknesses": "NO research, NO cost analysis, no architecture doc"},
    {"id": "38958", "teammates": "38981", "team": "103050", "estimated": "95%", "true_false": "21/22 (95%)", "grade": 94, "strengths": "HIGHEST SCORE! 500 tests, 1109 docstrings (96%), 19,411-word README, dual-interface, 3 CI/CD workflows, 28 screenshots, cost analysis", "weaknesses": "No Jupyter notebook, no pre-commit hooks"},
    {"id": "38959", "teammates": "38954, 38972", "team": "LLM_Agents_Tom_Igor_Roie", "estimated": "91%", "true_false": "20/22 (91%)", "grade": 86, "strengths": "144 tests, 413 docstrings, 3062-word README, complete prompt book, budget management, 14 screenshots, 55 commits", "weaknesses": "No research component, no pre-commit hooks"},
    {"id": "38960", "teammates": "38970", "team": "Group_6", "estimated": "82%", "true_false": "18/22 (82%)", "grade": 89, "strengths": "2 comprehensive Jupyter notebooks with plots/analysis, 173 tests, excellent documentation, complete PRD, cost analysis, 16 screenshots", "weaknesses": "No pre-commit hooks, could improve CI/CD coverage"},
    {"id": "38961", "teammates": "38964, 38985", "team": "The_Red_Team_Omry_Itamar", "estimated": "27%", "true_false": "6/22 (27%)", "grade": 44, "strengths": "Basic Flask implementation, 9 screenshots", "weaknesses": "NO tests, NO research, NO architecture docs, minimal documentation (5 docstrings), poor security, weak git"},
    {"id": "38962", "teammates": "38963, 38990", "team": "Oren_Danielle", "estimated": "50%", "true_false": "11/22 (50%)", "grade": 55, "strengths": "Basic FastAPI implementation, 11 screenshots, decent README structure", "weaknesses": "NO tests, NO research, NO cost analysis, minimal documentation (38 docstrings)"},
    {"id": "38963", "teammates": "38962", "team": "Oren_Danielle", "estimated": "50%", "true_false": "11/22 (50%)", "grade": 55, "strengths": "Basic FastAPI implementation, 11 screenshots, decent README structure", "weaknesses": "NO tests, NO research, NO cost analysis, minimal documentation (38 docstrings), weak git (4 commits)"},
    {"id": "38964", "teammates": "38961", "team": "The_Red_Team_Omry_Itamar", "estimated": "27%", "true_false": "6/22 (27%)", "grade": 44, "strengths": "Has basic Flask implementation with 9 screenshots", "weaknesses": "NO tests, NO research, NO architecture docs, minimal documentation (5 docstrings), poor security (secrets exposed), weak git (4 commits)"},
    {"id": "38965", "teammates": "38966, 38967", "team": "A_Star_Is_Born", "estimated": "64%", "true_false": "14/22 (64%)", "grade": 66, "strengths": "Good documentation (272 docstrings), 13 screenshots, decent README (1866 words)", "weaknesses": "NO research, NO cost analysis, minimal testing (12 tests), weak git"},
    {"id": "38966", "teammates": "38965, 38967", "team": "A_Star_Is_Born", "estimated": "64%", "true_false": "14/22 (64%)", "grade": 66, "strengths": "Good documentation (272 docstrings), 13 screenshots, decent README (1866 words)", "weaknesses": "NO research, NO cost analysis, minimal testing (12 tests), weak git (7 commits)"},
    {"id": "38967", "teammates": "38965, 38966", "team": "A_Star_Is_Born", "estimated": "DUPLICATE", "true_false": "14/22 (64%)", "grade": 66, "strengths": "DUPLICATE of 38966 - Good documentation, 13 screenshots", "weaknesses": "NO research, NO cost analysis, minimal testing"},
    {"id": "38968", "teammates": "38982, 38993", "team": "30af9fe9-f034-4e99-972f-319b8171fb1d", "estimated": "68%", "true_false": "15/22 (68%)", "grade": 71, "strengths": "Has Jupyter notebook with plots (rare!), 3 ADRs, good git with 4 feature branches, 259 docstrings", "weaknesses": "No cost analysis, minimal testing (23 tests), no CI/CD"},
    {"id": "38969", "teammates": "38971", "team": "BENSALMON", "estimated": "73%", "true_false": "16/22 (73%)", "grade": 82, "strengths": "Strong testing (212 tests, 93% coverage), 441 docstrings, 6029-word README, 21 screenshots, good git (33 commits)", "weaknesses": "No PRD, no research, no cost analysis, no pre-commit hooks"},
    {"id": "38970", "teammates": "38960", "team": "Group_6", "estimated": "DUPLICATE", "true_false": "18/22 (82%)", "grade": 89, "strengths": "DUPLICATE of 38960 - 2 Jupyter notebooks, 173 tests, excellent documentation", "weaknesses": "No pre-commit hooks, could improve CI/CD"},
    {"id": "38971", "teammates": "38969", "team": "BENSALMON", "estimated": "DUPLICATE", "true_false": "16/22 (73%)", "grade": 82, "strengths": "DUPLICATE of 38969 - 212 tests (93% coverage), 441 docstrings, 6029-word README", "weaknesses": "No PRD, no research, no cost analysis"},
    {"id": "38972", "teammates": "38949, 38950, 38953, 38954, 38959", "team": "The_Awesome_Foursome / LLM_Agents", "estimated": "DUPLICATE", "true_false": "19/22 (86%)", "grade": 87, "strengths": "DUPLICATE of 38953 - Complete PRD, 130 tests, research notebook, cost analysis", "weaknesses": "No pre-commit hooks"},
    {"id": "38973", "teammates": "38950", "team": "The_Awesome_Foursome", "estimated": "DUPLICATE", "true_false": "19/22 (86%)", "grade": 87, "strengths": "DUPLICATE of 38953 - Complete PRD, 130 tests, 654 docstrings, research notebook, cost analysis, 27 screenshots", "weaknesses": "No pre-commit hooks, could improve test coverage"},
    {"id": "38974", "teammates": "38955, 38956, 38957", "team": "The_Sucsess", "estimated": "64%", "true_false": "14/22 (64%)", "grade": 67, "strengths": "Good documentation (368 docstrings), decent testing (33 tests), good git (23 commits)", "weaknesses": "NO research, NO cost analysis, no architecture doc, no CI/CD"},
    {"id": "38975", "teammates": "38976, 38977, 38986, 38992", "team": "The_Surfers", "estimated": "59%", "true_false": "13/22 (59%)", "grade": 60, "strengths": "Perfect docstring coverage (160/160 = 100%), good security, decent git (22 commits)", "weaknesses": "NO research, NO cost analysis, minimal testing (21 tests), no architecture doc"},
    {"id": "38976", "teammates": "38975, 38977, 38986, 38992", "team": "The_Surfers", "estimated": "DUPLICATE", "true_false": "13/22 (59%)", "grade": 60, "strengths": "DUPLICATE of 38986 - Perfect docstring coverage (100%), good security", "weaknesses": "NO research, NO cost analysis, minimal testing"},
    {"id": "38977", "teammates": "38975, 38976", "team": "The_Surfers", "estimated": "DUPLICATE", "true_false": "13/22 (59%)", "grade": 60, "strengths": "DUPLICATE of 38986 - Perfect docstring coverage (160/160), good security, decent git history (22 commits)", "weaknesses": "NO research, NO cost analysis, minimal testing (21 tests), no architecture doc"},
    {"id": "38978", "teammates": "38979, 38980", "team": "almog_or", "estimated": "82%", "true_false": "18/22 (82%)", "grade": 88, "strengths": "10 Jupyter notebooks (exceptional research!), strong testing (91 tests), excellent documentation, comprehensive PRD", "weaknesses": "No cost analysis, no pre-commit hooks"},
    {"id": "38979", "teammates": "38978", "team": "almog_or", "estimated": "82%", "true_false": "18/22 (82%)", "grade": 88, "strengths": "10 Jupyter notebooks (exceptional research!), strong testing (91 tests), excellent documentation, comprehensive PRD", "weaknesses": "No cost analysis, no pre-commit hooks"},
    {"id": "38980", "teammates": "38978", "team": "almog_or", "estimated": "DUPLICATE", "true_false": "18/22 (82%)", "grade": 88, "strengths": "DUPLICATE of 38979 - 10 Jupyter notebooks, 91 tests, excellent documentation", "weaknesses": "No cost analysis, no pre-commit hooks"},
    {"id": "38981", "teammates": "38958", "team": "103050", "estimated": "95%", "true_false": "21/22 (95%)", "grade": 94, "strengths": "HIGHEST SCORE! 500 tests, 1109 docstrings (96%), 19,411-word README, dual-interface (Streamlit+Flask), 3 CI/CD workflows, 28 screenshots, complete cost analysis", "weaknesses": "No Jupyter notebook (only gap), no pre-commit hooks"},
    {"id": "38982", "teammates": "38968", "team": "30af9fe9-f034-4e99-972f-319b8171fb1d", "estimated": "68%", "true_false": "15/22 (68%)", "grade": 71, "strengths": "Has Jupyter notebook with plots (rare!), 3 ADRs, good git with 4 feature branches, 259 docstrings, complete prompt book", "weaknesses": "No cost analysis, minimal testing (23 tests), no CI/CD"},
    {"id": "38983", "teammates": "38984", "team": "Yossi-Yeuda-chat-bot", "estimated": "45%", "true_false": "10/22 (45%)", "grade": 55, "strengths": "Basic implementation with 11 screenshots, some documentation (142 docstrings)", "weaknesses": "NO research, NO cost analysis, NO tests, poor security, weak git"},
    {"id": "38984", "teammates": "38983", "team": "Yossi-Yeuda-chat-bot", "estimated": "45%", "true_false": "10/22 (45%)", "grade": 55, "strengths": "Basic implementation with 11 screenshots, some documentation (142 docstrings)", "weaknesses": "NO research, NO cost analysis, NO tests, poor security (hardcoded secrets), weak git (8 commits)"},
    {"id": "38985", "teammates": "38961", "team": "The_Red_Team_Omry_Itamar", "estimated": "DUPLICATE", "true_false": "6/22 (27%)", "grade": 44, "strengths": "DUPLICATE of 38964 - Basic Flask implementation, 9 screenshots", "weaknesses": "NO tests, NO research, NO architecture docs, minimal documentation"},
    {"id": "38986", "teammates": "38975, 38976", "team": "The_Surfers", "estimated": "27%", "true_false": "13/22 (59%)", "grade": 60, "strengths": "Perfect docstring coverage (160/160 = 100%), good security, decent git (22 commits), 7 screenshots", "weaknesses": "NO research, NO cost analysis, minimal testing (21 tests), no architecture doc"},
    {"id": "38988", "teammates": "38950", "team": "The_Awesome_Foursome", "estimated": "DUPLICATE", "true_false": "19/22 (86%)", "grade": 87, "strengths": "DUPLICATE of 38953 - Complete PRD, 130 tests, 654 docstrings, research notebook, cost analysis", "weaknesses": "No pre-commit hooks, could improve test coverage"},
    {"id": "38989", "teammates": "38950", "team": "The_Awesome_Foursome", "estimated": "DUPLICATE", "true_false": "19/22 (86%)", "grade": 87, "strengths": "DUPLICATE of 38953 - Complete PRD, 130 tests, 654 docstrings, research notebook, cost analysis", "weaknesses": "No pre-commit hooks, could improve test coverage"},
    {"id": "38990", "teammates": "38962", "team": "Oren_Danielle", "estimated": "DUPLICATE", "true_false": "11/22 (50%)", "grade": 55, "strengths": "DUPLICATE of 38963 - Basic FastAPI implementation, 11 screenshots", "weaknesses": "NO tests, NO research, NO cost analysis, minimal documentation"},
    {"id": "38992", "teammates": "38975, 38976", "team": "The_Surfers", "estimated": "DUPLICATE", "true_false": "13/22 (59%)", "grade": 60, "strengths": "DUPLICATE of 38986 - Perfect docstring coverage (160/160), good security, decent git", "weaknesses": "NO research, NO cost analysis, minimal testing (21 tests)"},
    {"id": "38993", "teammates": "38968", "team": "30af9fe9-f034-4e99-972f-319b8171fb1d", "estimated": "DUPLICATE", "true_false": "15/22 (68%)", "grade": 71, "strengths": "DUPLICATE of 38982 - Has Jupyter notebook with plots, 3 ADRs, good git with feature branches", "weaknesses": "No cost analysis, minimal testing, no CI/CD"},
    {"id": "59373", "teammates": "38951", "team": "Agents2025-2026", "estimated": "DUPLICATE", "true_false": "15/22 (68%)", "grade": 68, "strengths": "DUPLICATE of 38952 - Good documentation (382 docstrings), FastAPI implementation, 29 commits", "weaknesses": "NO research, NO cost analysis, minimal testing (8 tests)"},
    {"id": "59375", "teammates": "38954, 38972", "team": "LLM_Agents_Tom_Igor_Roie", "estimated": "DUPLICATE", "true_false": "20/22 (91%)", "grade": 86, "strengths": "DUPLICATE of 38959 - 144 tests, 413 docstrings, complete prompt book, budget management, 14 screenshots", "weaknesses": "No research component, no pre-commit hooks"},
    {"id": "59376", "teammates": "38971", "team": "BENSALMON", "estimated": "DUPLICATE", "true_false": "16/22 (73%)", "grade": 82, "strengths": "DUPLICATE of 38969 - 212 tests (93% coverage), 441 docstrings, 6029-word README, 21 screenshots", "weaknesses": "No PRD, no research, no cost analysis, no pre-commit hooks"},
    {"id": "59378", "teammates": "38958", "team": "103050", "estimated": "DUPLICATE", "true_false": "21/22 (95%)", "grade": 94, "strengths": "DUPLICATE of 38981 - HIGHEST SCORE! 500 tests, 1109 docstrings, 19,411-word README, dual-interface, 3 CI/CD workflows", "weaknesses": "No Jupyter notebook, no pre-commit hooks"},
]

# Sort by Student ID
all_students.sort(key=lambda x: int(x['id']))

# Write data
for row_num, student in enumerate(all_students, 2):
    ws.cell(row=row_num, column=1).value = student['id']
    ws.cell(row=row_num, column=1).alignment = center_align
    ws.cell(row=row_num, column=1).border = thin_border

    ws.cell(row=row_num, column=2).value = student['teammates']
    ws.cell(row=row_num, column=2).alignment = center_align
    ws.cell(row=row_num, column=2).border = thin_border

    ws.cell(row=row_num, column=3).value = student['team']
    ws.cell(row=row_num, column=3).alignment = left_align
    ws.cell(row=row_num, column=3).border = thin_border

    ws.cell(row=row_num, column=4).value = student['estimated']
    ws.cell(row=row_num, column=4).alignment = center_align
    ws.cell(row=row_num, column=4).border = thin_border

    ws.cell(row=row_num, column=5).value = student['true_false']
    ws.cell(row=row_num, column=5).alignment = center_align
    ws.cell(row=row_num, column=5).border = thin_border

    grade_cell = ws.cell(row=row_num, column=6)
    grade_cell.value = student['grade']
    grade_cell.alignment = center_align
    grade_cell.border = thin_border
    grade_cell.font = Font(bold=True, size=11)

    if student['grade'] >= 90:
        grade_cell.fill = grade_a
    elif student['grade'] >= 80:
        grade_cell.fill = grade_b
    elif student['grade'] >= 70:
        grade_cell.fill = grade_c
    elif student['grade'] >= 60:
        grade_cell.fill = grade_d
    else:
        grade_cell.fill = grade_f

    ws.cell(row=row_num, column=7).value = ""
    ws.cell(row=row_num, column=7).alignment = center_align
    ws.cell(row=row_num, column=7).border = thin_border
    ws.cell(row=row_num, column=7).fill = yellow_fill

    ws.cell(row=row_num, column=8).value = ""
    ws.cell(row=row_num, column=8).alignment = center_align
    ws.cell(row=row_num, column=8).border = thin_border
    ws.cell(row=row_num, column=8).fill = yellow_fill

    ws.cell(row=row_num, column=9).value = student['strengths']
    ws.cell(row=row_num, column=9).alignment = left_align
    ws.cell(row=row_num, column=9).border = thin_border

    ws.cell(row=row_num, column=10).value = student['weaknesses']
    ws.cell(row=row_num, column=10).alignment = left_align
    ws.cell(row=row_num, column=10).border = thin_border

# Freeze top row
ws.freeze_panes = "A2"

# Add summary statistics
summary_row = len(all_students) + 3
ws.cell(row=summary_row, column=1).value = "SUMMARY STATISTICS"
ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
ws.merge_cells(f'A{summary_row}:C{summary_row}')

stats = [
    ("Total Students:", len(all_students)),
    ("Average Grade:", f"=AVERAGE(F2:F{len(all_students)+1})"),
    ("Highest Grade:", f"=MAX(F2:F{len(all_students)+1})"),
    ("Lowest Grade:", f"=MIN(F2:F{len(all_students)+1})"),
    ("Pass Rate (>=60):", f"=COUNTIF(F2:F{len(all_students)+1},\">=60\")/COUNT(F2:F{len(all_students)+1})"),
    ("A Grades (>=90):", f"=COUNTIF(F2:F{len(all_students)+1},\">=90\")"),
    ("B Grades (80-89):", f"=COUNTIFS(F2:F{len(all_students)+1},\">=80\",F2:F{len(all_students)+1},\"<90\")"),
    ("C Grades (70-79):", f"=COUNTIFS(F2:F{len(all_students)+1},\">=70\",F2:F{len(all_students)+1},\"<80\")"),
    ("D Grades (60-69):", f"=COUNTIFS(F2:F{len(all_students)+1},\">=60\",F2:F{len(all_students)+1},\"<70\")"),
    ("F Grades (<60):", f"=COUNTIF(F2:F{len(all_students)+1},\"<60\")"),
]

for i, (label, value) in enumerate(stats, 1):
    stat_row = summary_row + i
    ws.cell(row=stat_row, column=1).value = label
    ws.cell(row=stat_row, column=1).font = Font(bold=True)
    ws.cell(row=stat_row, column=2).value = value
    if isinstance(value, str) and value.startswith("="):
        ws.cell(row=stat_row, column=2).number_format = '0.00' if "AVERAGE" in value or "RATE" in value else '0'

# Save workbook
output_path = r"C:\Users\Guest1\CoOp\Runi\Assignment3_Grading_Summary.xlsx"
wb.save(output_path)
print(f"Excel sheet created: {output_path}")
print(f"Total students: {len(all_students)}")
